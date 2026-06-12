#!/usr/bin/env python3
"""Build a daily token usage report and optionally send it to DingTalk.

The script is dry-run by default. Use --send only after the config maps usage
identities to DingTalk users and the target group has been verified.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import os
import subprocess
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from zoneinfo import ZoneInfo
except ImportError:  # pragma: no cover - Python < 3.9 fallback
    ZoneInfo = None  # type: ignore[assignment]


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_CONFIG = ROOT / "configs" / "token_usage_dingtalk.example.json"
LOCAL_ENV_PATH = ROOT / ".env"
LOCAL_ENV_KEYS = {
    "OPENAI_ADMIN_KEY",
    "OPENAI_API_KEY",
    "DINGTALK_WEBHOOK_TOKEN",
    "DINGTALK_ROBOT_CODE",
    "DINGTALK_GROUP_OPEN_CONVERSATION_ID",
}


class TokenUsageError(RuntimeError):
    def __init__(self, message: str, code: str = "token_usage_error") -> None:
        super().__init__(message)
        self.code = code


@dataclass
class UsageRecord:
    usage_id: str
    name: str | None = None
    input_tokens: int = 0
    output_tokens: int = 0
    amount: float = 0.0
    input_cached_tokens: int = 0
    input_audio_tokens: int = 0
    output_audio_tokens: int = 0
    requests: int = 0

    @property
    def total_tokens(self) -> int:
        return self.input_tokens + self.output_tokens

    def add(self, other: "UsageRecord") -> None:
        if not self.name and other.name:
            self.name = other.name
        self.input_tokens += int_value(other.input_tokens)
        self.output_tokens += int_value(other.output_tokens)
        self.amount = float_value(self.amount) + float_value(other.amount)
        self.input_cached_tokens += int_value(other.input_cached_tokens)
        self.input_audio_tokens += int_value(other.input_audio_tokens)
        self.output_audio_tokens += int_value(other.output_audio_tokens)
        self.requests += int_value(other.requests)

    def to_dict(self) -> dict[str, Any]:
        return {
            "usage_id": self.usage_id,
            "name": self.name,
            "input_tokens": self.input_tokens,
            "output_tokens": self.output_tokens,
            "total_tokens": self.total_tokens,
            "amount": round(float_value(self.amount), 4),
            "input_cached_tokens": self.input_cached_tokens,
            "input_audio_tokens": self.input_audio_tokens,
            "output_audio_tokens": self.output_audio_tokens,
            "requests": self.requests,
        }


def load_local_env(path: Path = LOCAL_ENV_PATH) -> None:
    if not path.exists():
        return
    try:
        lines = path.read_text(encoding="utf-8").splitlines()
    except OSError:
        return
    for raw_line in lines:
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if line.startswith("export "):
            line = line[len("export ") :].strip()
        if "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip().lstrip("\ufeff")
        if key not in LOCAL_ENV_KEYS or os.environ.get(key):
            continue
        value = value.strip()
        if len(value) >= 2 and value[0] == value[-1] and value[0] in {"'", '"'}:
            value = value[1:-1]
        os.environ[key] = value


def read_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)
    if not isinstance(payload, dict):
        raise TokenUsageError("config root must be an object", code="config_error")
    return payload


def resolve_path(value: str | Path, default_root: Path = ROOT) -> Path:
    path = Path(value).expanduser()
    return path if path.is_absolute() else default_root / path


def timezone_from_name(name: str) -> dt.tzinfo:
    if name in {"Asia/Shanghai", "Asia/Chongqing", "GMT+8", "UTC+8"}:
        return dt.timezone(dt.timedelta(hours=8), name="Asia/Shanghai")
    if name.upper() == "UTC":
        return dt.timezone.utc
    if ZoneInfo is not None:
        try:
            return ZoneInfo(name)
        except Exception:
            pass
    raise TokenUsageError(f"unsupported timezone: {name}", code="config_error")


def parse_report_window(config: dict[str, Any], date_override: str | None) -> dict[str, Any]:
    source = object_field(config, "source")
    tz = timezone_from_name(str(source.get("timezone") or config.get("timezone") or "Asia/Shanghai"))
    lookback_days = int(source.get("lookback_days", config.get("lookback_days", 1)))
    now = dt.datetime.now(tz)
    if date_override:
        report_date = dt.date.fromisoformat(date_override)
    else:
        report_date = (now - dt.timedelta(days=lookback_days)).date()
    start_local = dt.datetime.combine(report_date, dt.time.min, tzinfo=tz)
    end_local = start_local + dt.timedelta(days=1)
    return {
        "timezone": getattr(tz, "key", None) or str(tz),
        "report_date": report_date.isoformat(),
        "start_time": int(start_local.timestamp()),
        "end_time": int(end_local.timestamp()),
        "start_local": start_local.isoformat(),
        "end_local": end_local.isoformat(),
    }


def object_field(payload: dict[str, Any], key: str) -> dict[str, Any]:
    value = payload.get(key)
    if not isinstance(value, dict):
        raise TokenUsageError(f"config must include object field: {key}", code="config_error")
    return value


def list_field(payload: dict[str, Any], key: str) -> list[Any]:
    value = payload.get(key)
    if not isinstance(value, list):
        raise TokenUsageError(f"config must include list field: {key}", code="config_error")
    return value


def int_value(value: Any, default: int = 0) -> int:
    if value is None or value == "":
        return default
    try:
        return int(float(str(value).replace(",", "")))
    except (TypeError, ValueError):
        return default


def float_value(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    try:
        return float(str(value).replace(",", ""))
    except (TypeError, ValueError):
        return default


def bool_value(value: Any, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    if value is None or value == "":
        return default
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "y", "on"}:
        return True
    if text in {"0", "false", "no", "n", "off"}:
        return False
    return default


def normalize_group_by(value: Any) -> list[str]:
    if isinstance(value, str):
        return [item.strip() for item in value.split(",") if item.strip()]
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    return ["user_id"]


def choose_identity_field(source: dict[str, Any]) -> str:
    identity_field = str(source.get("identity_field") or "").strip()
    if identity_field:
        return identity_field
    group_by = normalize_group_by(source.get("group_by", ["user_id"]))
    for candidate in ("user_id", "api_key_id", "project_id", "model"):
        if candidate in group_by:
            return candidate
    return group_by[0] if group_by else "user_id"


def fetch_openai_usage(source: dict[str, Any], window: dict[str, Any]) -> dict[str, UsageRecord]:
    api_key_env = str(source.get("api_key_env") or "OPENAI_ADMIN_KEY")
    api_key = os.environ.get(api_key_env) or os.environ.get("OPENAI_API_KEY")
    if not api_key:
        raise TokenUsageError(
            f"missing OpenAI API key environment variable: {api_key_env}",
            code="credentials_missing",
        )

    base_url = str(source.get("base_url") or "https://api.openai.com/v1").rstrip("/")
    usage_path = str(source.get("usage_path") or "/organization/usage/completions")
    url = f"{base_url}/{usage_path.lstrip('/')}"
    group_by = normalize_group_by(source.get("group_by", ["user_id"]))
    identity_field = choose_identity_field(source)
    limit = int(source.get("limit", 31))
    max_pages = int(source.get("max_pages", 20))
    bucket_width = str(source.get("bucket_width") or "1d")

    params: dict[str, Any] = {
        "start_time": int(window["start_time"]),
        "end_time": int(window["end_time"]),
        "bucket_width": bucket_width,
        "limit": limit,
        "group_by": group_by,
    }
    for optional_key in ("project_ids", "user_ids", "api_key_ids", "models"):
        if source.get(optional_key):
            params[optional_key] = source[optional_key]
    if "batch" in source:
        params["batch"] = bool(source["batch"])

    records: dict[str, UsageRecord] = {}
    page: str | None = None
    for _ in range(max_pages):
        request_params = dict(params)
        if page:
            request_params["page"] = page
        query = urllib.parse.urlencode(request_params, doseq=True)
        req = urllib.request.Request(
            f"{url}?{query}",
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            },
        )
        try:
            with urllib.request.urlopen(req, timeout=int(source.get("timeout_sec", 60))) as response:
                payload = json.loads(response.read().decode("utf-8"))
        except urllib.error.HTTPError as error:
            body = error.read().decode("utf-8", "replace")
            raise TokenUsageError(f"OpenAI usage API failed: HTTP {error.code}: {body[:500]}", code="source_error")
        except (urllib.error.URLError, TimeoutError, OSError) as error:
            raise TokenUsageError(f"OpenAI usage API request failed: {error}", code="source_error")

        for bucket in payload.get("data") or []:
            if not isinstance(bucket, dict):
                continue
            for result in bucket.get("results") or []:
                if not isinstance(result, dict):
                    continue
                usage_id = str(result.get(identity_field) or "UNMAPPED")
                record = UsageRecord(
                    usage_id=usage_id,
                    input_tokens=int_value(result.get("input_tokens")),
                    output_tokens=int_value(result.get("output_tokens")),
                    input_cached_tokens=int_value(result.get("input_cached_tokens")),
                    input_audio_tokens=int_value(result.get("input_audio_tokens")),
                    output_audio_tokens=int_value(result.get("output_audio_tokens")),
                    requests=int_value(result.get("num_model_requests")),
                )
                records.setdefault(usage_id, UsageRecord(usage_id)).add(record)
        page = payload.get("next_page")
        if not page:
            break
    else:
        raise TokenUsageError("OpenAI usage API pagination exceeded max_pages", code="source_error")

    return records


def fetch_csv_usage(source: dict[str, Any], window: dict[str, Any]) -> dict[str, UsageRecord]:
    path = resolve_path(str(source.get("path") or "inputs/token_usage_daily.csv"))
    if not path.exists():
        raise TokenUsageError(f"CSV source file does not exist: {path}", code="source_error")
    id_column = str(source.get("id_column") or "usage_id")
    date_column = str(source.get("date_column") or "date")
    input_column = str(source.get("input_tokens_column") or "input_tokens")
    output_column = str(source.get("output_tokens_column") or "output_tokens")
    total_column = str(source.get("total_tokens_column") or "total_tokens")
    requests_column = str(source.get("requests_column") or "requests")
    report_date = str(window["report_date"])
    records: dict[str, UsageRecord] = {}

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            if date_column in row and row.get(date_column) and str(row.get(date_column)).strip() != report_date:
                continue
            usage_id = str(row.get(id_column) or "").strip()
            if not usage_id:
                continue
            input_tokens = int_value(row.get(input_column))
            output_tokens = int_value(row.get(output_column))
            total_tokens = int_value(row.get(total_column))
            if total_tokens and not input_tokens and not output_tokens:
                input_tokens = total_tokens
            record = UsageRecord(
                usage_id=usage_id,
                name=str(row.get(source.get("name_column", "")) or usage_id).strip(),
                input_tokens=input_tokens,
                output_tokens=output_tokens,
                requests=int_value(row.get(requests_column)),
            )
            records.setdefault(usage_id, UsageRecord(usage_id)).add(record)
    return records


def normalize_column_name(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "").replace("_", "").replace("-", "")


def configured_column(source: dict[str, Any], key: str, aliases: list[str]) -> list[str]:
    value = source.get(key)
    candidates: list[str] = []
    if isinstance(value, str) and value.strip():
        candidates.append(value.strip())
    configured_aliases = source.get("column_aliases")
    if isinstance(configured_aliases, dict):
        extra = configured_aliases.get(key)
        if isinstance(extra, str):
            candidates.append(extra)
        elif isinstance(extra, list):
            candidates.extend(str(item) for item in extra if str(item).strip())
    candidates.extend(aliases)
    return candidates


def find_column(headers: list[Any], candidates: list[str], required: bool = False) -> str | None:
    normalized = {normalize_column_name(header): str(header).strip() for header in headers if header not in (None, "")}
    for candidate in candidates:
        direct = normalized.get(normalize_column_name(candidate))
        if direct:
            return direct
    for header in headers:
        text = str(header or "").strip()
        normalized_text = normalize_column_name(text)
        if not normalized_text:
            continue
        for candidate in candidates:
            normalized_candidate = normalize_column_name(candidate)
            if normalized_candidate and normalized_candidate in normalized_text:
                return text
    if required:
        raise TokenUsageError(f"cannot find required column, candidates={candidates}", code="source_error")
    return None


def value_as_date(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    for fmt, width in (("%Y-%m-%d", 10), ("%Y/%m/%d", 10), ("%Y.%m.%d", 10), ("%Y%m%d", 8)):
        try:
            return dt.datetime.strptime(text[:width], fmt).date().isoformat()
        except ValueError:
            pass
    return text[:10]


def detect_header_row(rows: list[list[Any]], source: dict[str, Any]) -> int:
    if source.get("header_row"):
        return int(source["header_row"])
    token_candidates = configured_column(
        source,
        "total_tokens_column",
        ["total_tokens", "总tokens", "总token", "token消耗", "tokens消耗", "消耗token", "消耗tokens", "总消耗", "token", "tokens"],
    )
    name_candidates = configured_column(
        source,
        "name_column",
        ["姓名", "员工姓名", "员工", "使用人", "用户", "成员", "人员", "发起人", "name", "employee", "member"],
    )
    id_candidates = configured_column(source, "id_column", ["usage_id", "user_id", "用户id", "员工id", "账号", "邮箱", "工号", "id", "account"])
    for index, row in enumerate(rows[:20], start=1):
        headers = [cell for cell in row if cell not in (None, "")]
        if len(headers) < 2:
            continue
        has_token = find_column(headers, token_candidates, required=False) is not None
        has_person = (
            find_column(headers, name_candidates, required=False) is not None
            or find_column(headers, id_candidates, required=False) is not None
        )
        if has_token and has_person:
            return index
    raise TokenUsageError("cannot detect Excel header row; set source.header_row and column names", code="source_error")


def worksheet_rows(path: Path, sheet_name: str | None) -> tuple[list[list[Any]], str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise TokenUsageError(f"Excel sheet not found: {sheet_name}", code="source_error")
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook[workbook.sheetnames[0]]
    return [list(row) for row in worksheet.iter_rows(values_only=True)], worksheet.title


def fetch_xlsx_usage(source: dict[str, Any], window: dict[str, Any]) -> dict[str, UsageRecord]:
    path = resolve_path(str(source.get("path") or "inputs/token_usage_daily.xlsx"))
    if not path.exists():
        raise TokenUsageError(f"Excel source file does not exist: {path}", code="source_error")
    rows, _sheet = worksheet_rows(path, str(source.get("sheet") or "") or None)
    if not rows:
        return {}
    if source.get("format") == "smartbi_pivot":
        return fetch_smartbi_pivot_usage(source, rows)
    header_row = detect_header_row(rows, source)
    headers = [str(cell).strip() if cell not in (None, "") else "" for cell in rows[header_row - 1]]
    name_column = find_column(
        headers,
        configured_column(source, "name_column", ["姓名", "员工姓名", "员工", "使用人", "用户", "成员", "人员", "发起人", "name", "employee", "member"]),
    )
    id_column = find_column(headers, configured_column(source, "id_column", ["usage_id", "user_id", "用户id", "员工id", "账号", "邮箱", "工号", "id", "account"]))
    total_column = find_column(
        headers,
        configured_column(
            source,
            "total_tokens_column",
            ["total_tokens", "总tokens", "总token", "token消耗", "tokens消耗", "消耗token", "消耗tokens", "总消耗", "token", "tokens"],
        ),
        required=True,
    )
    input_column = find_column(headers, configured_column(source, "input_tokens_column", ["input_tokens", "输入tokens", "输入token"]))
    output_column = find_column(headers, configured_column(source, "output_tokens_column", ["output_tokens", "输出tokens", "输出token"]))
    requests_column = find_column(headers, configured_column(source, "requests_column", ["requests", "请求次数", "调用次数", "次数"]))
    date_column = find_column(headers, configured_column(source, "date_column", ["date", "日期", "统计日期", "使用日期"]))
    if not id_column and not name_column:
        raise TokenUsageError("Excel source needs id_column or name_column", code="source_error")

    records: dict[str, UsageRecord] = {}
    report_date = str(window["report_date"])
    header_index = {header: index for index, header in enumerate(headers) if header}
    for raw_row in rows[header_row:]:
        row = {header: raw_row[index] if index < len(raw_row) else None for header, index in header_index.items()}
        if date_column and value_as_date(row.get(date_column)) != report_date:
            continue
        name = str(row.get(name_column) or "").strip() if name_column else ""
        usage_id = str(row.get(id_column) or name).strip() if id_column else name
        if not usage_id:
            continue
        input_tokens = int_value(row.get(input_column)) if input_column else 0
        output_tokens = int_value(row.get(output_column)) if output_column else 0
        total_tokens = int_value(row.get(total_column))
        if total_tokens and not input_tokens and not output_tokens:
            input_tokens = total_tokens
        record = UsageRecord(
            usage_id=usage_id,
            name=name or usage_id,
            input_tokens=input_tokens,
            output_tokens=output_tokens,
            requests=int_value(row.get(requests_column)) if requests_column else 0,
        )
        records.setdefault(usage_id, UsageRecord(usage_id, name=name or usage_id)).add(record)
    return records


def one_based_index(source: dict[str, Any], key: str, default: int) -> int:
    value = source.get(key, default)
    return int(value) - 1


def fetch_smartbi_pivot_usage(source: dict[str, Any], rows: list[list[Any]]) -> dict[str, UsageRecord]:
    data_start_row = int(source.get("data_start_row") or 7) - 1
    name_index = one_based_index(source, "name_column_index", 4)
    total_index = one_based_index(source, "total_tokens_column_index", 5)
    requests_index = one_based_index(source, "requests_column_index", 6)
    amount_index = one_based_index(source, "amount_column_index", 7)
    token_multiplier = int(source.get("token_multiplier") or 1)
    records: dict[str, UsageRecord] = {}

    for raw_row in rows[data_start_row:]:
        name = str(raw_row[name_index] if name_index < len(raw_row) else "" or "").strip()
        if not name:
            continue
        raw_total = raw_row[total_index] if total_index < len(raw_row) else 0
        try:
            total_tokens = int(round(float(str(raw_total).replace(",", "")) * token_multiplier))
        except (TypeError, ValueError):
            total_tokens = 0
        record = UsageRecord(
            usage_id=name,
            name=name,
            input_tokens=total_tokens,
            amount=float_value(raw_row[amount_index] if amount_index < len(raw_row) else 0),
            requests=int_value(raw_row[requests_index] if requests_index < len(raw_row) else 0),
        )
        records.setdefault(name, UsageRecord(name, name=name)).add(record)
    return records


def fetch_smartbi_task_usage(source: dict[str, Any], window: dict[str, Any], task_override: str | None = None) -> dict[str, UsageRecord]:
    smartbi_config = resolve_path(str(source.get("smartbi_config") or "configs/token_usage_smartbi_tasks.local.json"))
    task = str(task_override or source.get("task") or "")
    if not task:
        raise TokenUsageError("smartbi_task source requires source.task", code="config_error")
    command = [
        sys.executable,
        str(ROOT / "scripts" / "smartbi_cli.py"),
        "run",
        "--config",
        str(smartbi_config),
        "--task",
        task,
        "--run-date",
        str(window["report_date"]),
        "--overwrite",
        "--json",
    ]
    completed = subprocess.run(
        command,
        cwd=ROOT,
        text=True,
        capture_output=True,
        check=False,
        timeout=int(source.get("timeout_sec", 300)),
    )
    if completed.returncode != 0:
        raise TokenUsageError(f"SmartBI export failed: {completed.stderr.strip()}", code="source_error")
    try:
        payload = json.loads(completed.stdout)
    except json.JSONDecodeError as error:
        raise TokenUsageError(f"SmartBI export returned non-JSON output: {error}", code="source_error")
    output = payload.get("output")
    if not output:
        raise TokenUsageError("SmartBI export did not return output path", code="source_error")
    parse_config = dict(source.get("parse") if isinstance(source.get("parse"), dict) else {})
    parse_config.update({"type": "xlsx", "path": output})
    return fetch_xlsx_usage(parse_config, window)


def fetch_usage(config: dict[str, Any], window: dict[str, Any]) -> tuple[dict[str, UsageRecord], str]:
    source = object_field(config, "source")
    source_type = str(source.get("type") or "openai_usage_api")
    if source_type == "openai_usage_api":
        return fetch_openai_usage(source, window), choose_identity_field(source)
    if source_type == "csv":
        return fetch_csv_usage(source, window), str(source.get("id_column") or "usage_id")
    if source_type == "xlsx":
        return fetch_xlsx_usage(source, window), str(source.get("id_column") or source.get("name_column") or "usage_id")
    if source_type == "smartbi_task":
        parse = source.get("parse") if isinstance(source.get("parse"), dict) else {}
        identity_field = str(parse.get("id_column") or parse.get("name_column") or "usage_id")
        return fetch_smartbi_task_usage(source, window), identity_field
    raise TokenUsageError(f"unsupported source.type: {source_type}", code="config_error")


def member_usage_id(member: dict[str, Any], identity_field: str) -> str:
    for key in ("usage_id", identity_field, "openai_user_id", "api_key_id", "project_id"):
        value = member.get(key)
        if value:
            return str(value)
    return ""


def build_report(
    config: dict[str, Any],
    usage: dict[str, UsageRecord],
    identity_field: str,
    window: dict[str, Any],
    daily_usage: dict[str, UsageRecord] | None = None,
) -> dict[str, Any]:
    thresholds = object_field(config, "thresholds")
    metric = str(thresholds.get("metric") or "tokens")
    min_total_tokens = int(thresholds.get("min_total_tokens", 0))
    min_amount = float_value(thresholds.get("min_amount", 0))
    include_zero_usage = bool(thresholds.get("include_zero_usage", True))
    raw_members = config.get("members", [])
    if raw_members is None:
        raw_members = []
    if not isinstance(raw_members, list):
        raise TokenUsageError("config.members must be a list when present", code="config_error")
    members = [item for item in raw_members if isinstance(item, dict) and item.get("enabled", True)]
    if not members:
        members = [
            {"name": record.name or usage_id, "usage_id": usage_id, "enabled": True}
            for usage_id, record in sorted(usage.items())
            if usage_id != "UNMAPPED"
        ]

    member_rows: list[dict[str, Any]] = []
    daily_usage = daily_usage or {}
    mapped_usage_ids: set[str] = set()
    for member in members:
        usage_id = member_usage_id(member, identity_field)
        if not usage_id:
            raise TokenUsageError(f"member is missing usage identity: {member}", code="config_error")
        mapped_usage_ids.add(usage_id)
        record = usage.get(usage_id, UsageRecord(usage_id))
        has_usage = usage_id in usage
        daily_record = daily_usage.get(usage_id, UsageRecord(usage_id))
        daily_amount = float_value(daily_record.amount)
        has_daily_usage = usage_id in daily_usage
        is_daily_zero = daily_amount == 0
        metric_value = float_value(record.amount) if metric == "amount" else float(record.total_tokens)
        threshold_value = min_amount if metric == "amount" else float(min_total_tokens)
        is_low = metric_value < threshold_value and (include_zero_usage or has_usage)
        member_rows.append(
            {
                "name": str(member.get("name") or record.name or usage_id),
                "usage_id": usage_id,
                "dingtalk_user_id": member.get("dingtalk_user_id"),
                "open_dingtalk_id": member.get("open_dingtalk_id"),
                "mobile": member.get("mobile"),
                "input_tokens": record.input_tokens,
                "output_tokens": record.output_tokens,
                "total_tokens": record.total_tokens,
                "amount": round(float_value(record.amount), 4),
                "requests": record.requests,
                "has_usage": has_usage,
                "daily_amount": round(daily_amount, 4),
                "daily_requests": daily_record.requests,
                "has_daily_usage": has_daily_usage,
                "is_daily_zero": is_daily_zero,
                "is_low": is_low,
            }
        )

    unmatched = [
        record.to_dict()
        for usage_id, record in sorted(usage.items())
        if usage_id not in mapped_usage_ids and usage_id != "UNMAPPED"
    ]
    low_consumers = [row for row in member_rows if row["is_low"]]
    daily_zero_consumers = [row for row in member_rows if row["is_daily_zero"]]
    total_tokens = sum(row["total_tokens"] for row in member_rows) + sum(item["total_tokens"] for item in unmatched)
    total_amount = sum(float(row["amount"]) for row in member_rows)
    active_members = sum(1 for row in member_rows if row["total_tokens"] > 0)

    return {
        "status": "ok",
        "report_date": window["report_date"],
        "window": window,
        "identity_field": identity_field,
        "thresholds": {
            "min_total_tokens": min_total_tokens,
            "min_amount": min_amount,
            "metric": metric,
            "include_zero_usage": include_zero_usage,
        },
        "counts": {
            "members": len(member_rows),
            "active_members": active_members,
            "low_consumers": len(low_consumers),
            "daily_zero_consumers": len(daily_zero_consumers),
            "unmatched_usage_ids": len(unmatched),
        },
        "total_tokens": total_tokens,
        "total_amount": round(total_amount, 4),
        "members": sorted(member_rows, key=lambda item: (float(item["amount"]), item["name"])),
        "low_consumers": sorted(low_consumers, key=lambda item: (float(item["amount"]), item["name"])),
        "daily_zero_consumers": sorted(daily_zero_consumers, key=lambda item: item["name"]),
        "unmatched_usage": unmatched,
    }


def at_marker(row: dict[str, Any], mode: str) -> str:
    if mode == "current_user" and row.get("open_dingtalk_id"):
        return f"<@{row['open_dingtalk_id']}>"
    if mode == "bot":
        if row.get("dingtalk_user_id"):
            return f"@{row['dingtalk_user_id']}"
        if row.get("open_dingtalk_id"):
            return f"@{row['open_dingtalk_id']}"
    if mode == "webhook":
        if row.get("mobile"):
            return f"@{row['mobile']}"
        if row.get("dingtalk_user_id"):
            return f"@{row['dingtalk_user_id']}"
    return f"@{row['name']}"


def format_number(value: int) -> str:
    return f"{value:,}"


def format_amount(value: float) -> str:
    text = f"{value:,.2f}"
    return text.rstrip("0").rstrip(".")


def build_markdown(report: dict[str, Any], config: dict[str, Any]) -> tuple[str, dict[str, list[str]]]:
    dingtalk = object_field(config, "dingtalk")
    mode = str(dingtalk.get("mode") or "current_user")
    max_rows = int(dingtalk.get("max_detail_rows", 30))
    report_date = report["report_date"]
    amount_threshold = float(report["thresholds"].get("min_amount") or 0)
    counts = report["counts"]
    at: dict[str, list[str]] = {"open_dingtalk_ids": [], "user_ids": [], "mobiles": []}

    lines = [
        f"## 每日 AI 金额消耗提醒（截至 {report_date}）",
        "",
        f"- 监督人数：{counts['members']} 人；金额低于 {format_amount(amount_threshold)} / 未匹配：{counts['low_consumers']} 人",
        f"- 当日 0 消耗：{counts.get('daily_zero_consumers', 0)} 人",
        f"- 总金额：{format_amount(float(report.get('total_amount') or 0))}",
        "",
    ]

    low_consumers = report["low_consumers"][:max_rows]
    if low_consumers:
        lines.append("### 点名提醒")
        lines.append("")
        for row in low_consumers:
            if not row.get("has_usage"):
                lines.append(f"- {row['name']}：未匹配")
            else:
                lines.append(f"- {row['name']}：金额 {format_amount(float(row['amount']))}")
            if row.get("open_dingtalk_id"):
                at["open_dingtalk_ids"].append(str(row["open_dingtalk_id"]))
            if row.get("dingtalk_user_id"):
                at["user_ids"].append(str(row["dingtalk_user_id"]))
            if row.get("mobile"):
                at["mobiles"].append(str(row["mobile"]))
    else:
        lines.append("今日没有低消耗人员。")

    daily_zero_consumers = report.get("daily_zero_consumers", [])[:max_rows]
    if daily_zero_consumers:
        lines.append("")
        lines.append("### 当日 0 消耗")
        lines.append("")
        for row in daily_zero_consumers:
            lines.append(f"- {row['name']}")

    return "\n".join(lines), at


def image_font(size: int, bold: bool = False) -> Any:
    font_names = [
        "msyhbd.ttc" if bold else "msyh.ttc",
        "simhei.ttf" if bold else "simsun.ttc",
        "arialbd.ttf" if bold else "arial.ttf",
    ]
    for name in font_names:
        path = Path("C:/Windows/Fonts") / name
        if path.exists():
            try:
                from PIL import ImageFont

                return ImageFont.truetype(str(path), size=size)
            except Exception:
                continue
    from PIL import ImageFont

    return ImageFont.load_default()


def draw_text_fit(draw: Any, xy: tuple[int, int], text: str, font: Any, fill: str, max_width: int) -> None:
    value = text
    while value:
        bbox = draw.textbbox((0, 0), value, font=font)
        if bbox[2] - bbox[0] <= max_width:
            break
        value = value[:-2] + "…"
    draw.text(xy, value or text[:1], font=font, fill=fill)


def render_report_image(report: dict[str, Any], config: dict[str, Any], path: Path) -> None:
    try:
        from PIL import Image, ImageDraw
    except Exception as error:
        raise TokenUsageError(f"Pillow is required to render notification image: {error}", code="image_error") from error

    dingtalk = object_field(config, "dingtalk")
    max_rows = int(dingtalk.get("max_detail_rows", 30))
    low_rows = list(report.get("low_consumers", []))[:max_rows]
    daily_rows = list(report.get("daily_zero_consumers", []))[:max_rows]
    counts = report["counts"]
    report_date = str(report["report_date"])
    amount_threshold = format_amount(float(report["thresholds"].get("min_amount") or 0))
    total_amount = format_amount(float(report.get("total_amount") or 0))

    width = 1080
    height = 1280
    margin = 36
    inner = 72
    card_left = margin
    card_top = 36
    card_right = width - margin
    card_bottom = height - 36
    image = Image.new("RGB", (width, height), "#F3EDE4")
    draw = ImageDraw.Draw(image)

    def text_width(text: str, font: Any) -> int:
        bbox = draw.textbbox((0, 0), text, font=font)
        return bbox[2] - bbox[0]

    def monthly_label(row: dict[str, Any]) -> tuple[str, bool]:
        name = str(row.get("name") or "")
        if not row.get("has_usage"):
            return f"{name} 未匹配", False
        return f"{name} 金额 {format_amount(float(row.get('amount') or 0))}", True

    title_font = image_font(50, bold=True)
    brand_font = image_font(24, bold=True)
    subtitle_font = image_font(26)
    metric_label_font = image_font(24)
    metric_value_font = image_font(36, bold=True)
    section_title_font = image_font(35, bold=True)
    section_subtitle_font = image_font(24)
    row_font = image_font(22, bold=True)
    row_value_font = image_font(22, bold=True)
    daily_font = image_font(25, bold=True)

    colors = {
        "ink": "#2A211B",
        "muted": "#7C6F65",
        "line": "#E8D7C6",
        "panel": "#FFFDF8",
        "soft": "#F4EFE8",
        "red": "#8A1F22",
        "amber": "#A56600",
        "teal": "#0F7B72",
    }

    draw.rounded_rectangle((card_left, card_top, card_right, card_bottom), radius=34, fill=colors["panel"], outline=colors["line"], width=1)
    x0 = card_left + inner - 36
    y0 = card_top + 40
    draw.text((x0, y0), "VIP-THINK", font=brand_font, fill=colors["red"])
    draw.text((x0, y0 + 48), "AI 消耗行动提醒", font=title_font, fill=colors["ink"])
    draw.text((x0, y0 + 108), f"{report_date} · 监督名单内金额总计 {total_amount}", font=subtitle_font, fill=colors["muted"])

    metric_y = y0 + 164
    metric_w = 206
    metric_gap = 47
    metrics = [
        ("监督人数", f"{counts['members']} 人", colors["teal"]),
        ("需关注", f"{counts['low_consumers']} 人", colors["amber"]),
        ("当日0消耗", f"{counts.get('daily_zero_consumers', 0)} 人", colors["red"]),
        ("阈值", amount_threshold, colors["ink"]),
    ]
    for index, (label, value, fill) in enumerate(metrics):
        x = x0 + index * (metric_w + metric_gap)
        draw.rounded_rectangle((x, metric_y, x + metric_w, metric_y + 110), radius=20, fill="#F8F4EF", outline=colors["line"], width=1)
        draw.text((x + 22, metric_y + 24), label, font=metric_label_font, fill=colors["muted"])
        draw_text_fit(draw, (x + 22, metric_y + 60), value, metric_value_font, fill, metric_w - 44)

    def draw_monthly_section(y: int) -> int:
        section_x = x0
        section_w = card_right - x0 - 36
        draw.rounded_rectangle((section_x, y, section_x + section_w, y + 372), radius=24, fill="#FFFFFF", outline=colors["line"], width=1)
        draw.text((section_x + 30, y + 32), "月累计点名", font=section_title_font, fill=colors["ink"])
        draw.text((section_x + 30, y + 74), "金额低于阈值或未匹配", font=section_subtitle_font, fill=colors["muted"])
        left_x = section_x + 30
        right_x = section_x + section_w // 2 + 12
        row_y = y + 112
        row_h = 48
        col_w = section_w // 2 - 54
        for index, row in enumerate(low_rows[:10]):
            col_x = left_x if index % 2 == 0 else right_x
            item_y = row_y + (index // 2) * row_h
            label, has_amount = monthly_label(row)
            if " " in label:
                name, value = label.split(" ", 1)
            else:
                name, value = label, ""
            draw.rounded_rectangle((col_x, item_y, col_x + col_w, item_y + 38), radius=12, fill=colors["soft"])
            draw_text_fit(draw, (col_x + 14, item_y + 7), name, row_font, colors["ink"], col_w - 160)
            value_color = colors["amber"] if has_amount else colors["red"]
            draw_text_fit(draw, (col_x + col_w - 134, item_y + 7), value, row_value_font, value_color, 120)
        return y + 400

    def draw_daily_section(y: int) -> int:
        section_x = x0
        section_w = card_right - x0 - 36
        draw.rounded_rectangle((section_x, y, section_x + section_w, y + 306), radius=24, fill="#FFFFFF", outline=colors["line"], width=1)
        draw.text((section_x + 30, y + 32), "当日 0 消耗", font=section_title_font, fill=colors["ink"])
        draw.text((section_x + 30, y + 74), "当天日维度没有金额记录", font=section_subtitle_font, fill=colors["muted"])
        chip_w = 206
        chip_h = 40
        gap_x = 18
        gap_y = 14
        start_x = section_x + 30
        start_y = y + 112
        for index, row in enumerate(daily_rows[:12]):
            col = index % 4
            line = index // 4
            chip_x = start_x + col * (chip_w + gap_x)
            chip_y = start_y + line * (chip_h + gap_y)
            draw.rounded_rectangle((chip_x, chip_y, chip_x + chip_w, chip_y + chip_h), radius=20, fill="#F1E9E0")
            draw_text_fit(draw, (chip_x + 18, chip_y + 6), str(row.get("name") or ""), daily_font, colors["ink"], chip_w - 36)
        return y + 334

    y = metric_y + 166
    y = draw_monthly_section(y)
    y = draw_daily_section(y + 2)

    path.parent.mkdir(parents=True, exist_ok=True)
    image.save(path, format="PNG", optimize=True)


def should_send(report: dict[str, Any], config: dict[str, Any]) -> bool:
    dingtalk = object_field(config, "dingtalk")
    if report["counts"]["low_consumers"] > 0:
        return True
    return bool(dingtalk.get("send_when_all_ok", True))


def env_or_config(dingtalk: dict[str, Any], key: str, env_key: str) -> str:
    value = dingtalk.get(key) or os.environ.get(env_key) or ""
    return str(value)


def run_dws_json(command: list[str], timeout_sec: int) -> dict[str, Any]:
    completed = subprocess.run(
        command,
        cwd=ROOT,
        text=True,
        encoding="utf-8",
        errors="replace",
        capture_output=True,
        check=False,
        timeout=timeout_sec,
    )
    payload: dict[str, Any] | None = None
    if completed.stdout.strip():
        try:
            payload = json.loads(completed.stdout)
        except json.JSONDecodeError:
            payload = None
    return {
        "returncode": completed.returncode,
        "command": command,
        "stdout": completed.stdout.strip(),
        "stderr": completed.stderr.strip(),
        "payload": payload,
    }


def dws_success(result: dict[str, Any]) -> bool:
    payload = result.get("payload")
    return result.get("returncode") == 0 and isinstance(payload, dict) and bool(payload.get("success", True))


def payload_result(result: dict[str, Any]) -> dict[str, Any]:
    payload = result.get("payload")
    if isinstance(payload, dict) and isinstance(payload.get("result"), dict):
        return payload["result"]
    return {}


def get_conversation_space_id(dws_command: str, group: str, timeout_sec: int) -> tuple[str | None, dict[str, Any]]:
    result = run_dws_json(
        [dws_command, "chat", "conversation-info", "--group", group, "--format", "json"],
        timeout_sec,
    )
    info = payload_result(result).get("conversationInfo")
    space_id = None
    if isinstance(info, dict):
        extension = info.get("extension")
        if isinstance(extension, dict):
            space_id = str(extension.get("newCSpaceIdIM") or "").strip() or None
    return space_id, result


def upload_dingtalk_image(
    dws_command: str,
    image_path: Path,
    group: str,
    timeout_sec: int,
) -> dict[str, Any]:
    space_id, conversation_result = get_conversation_space_id(dws_command, group, timeout_sec)
    upload_result: dict[str, Any] | None = None
    download_result: dict[str, Any] | None = None
    download_url = ""
    file_id = ""
    if not space_id:
        return {
            "ok": False,
            "error": "missing conversation space id",
            "conversation_info": conversation_result,
        }

    upload_result = run_dws_json(
        [
            dws_command,
            "drive",
            "upload",
            "--file",
            str(image_path),
            "--file-name",
            image_path.name,
            "--space-id",
            space_id,
            "--mime-type",
            "image/png",
            "--format",
            "json",
        ],
        timeout_sec,
    )
    if not dws_success(upload_result):
        return {
            "ok": False,
            "error": "drive upload failed",
            "space_id": space_id,
            "conversation_info": conversation_result,
            "upload": upload_result,
        }
    file_id = str(payload_result(upload_result).get("fileId") or "")
    if not file_id:
        return {
            "ok": False,
            "error": "drive upload did not return fileId",
            "space_id": space_id,
            "conversation_info": conversation_result,
            "upload": upload_result,
        }

    download_result = run_dws_json(
        [
            dws_command,
            "drive",
            "download",
            "--file-id",
            file_id,
            "--space-id",
            space_id,
            "--format",
            "json",
        ],
        timeout_sec,
    )
    if not dws_success(download_result):
        return {
            "ok": False,
            "error": "drive download url failed",
            "space_id": space_id,
            "file_id": file_id,
            "conversation_info": conversation_result,
            "upload": upload_result,
            "download": download_result,
        }
    download_url = str(payload_result(download_result).get("downloadUrl") or "")
    if not download_url:
        return {
            "ok": False,
            "error": "drive download did not return downloadUrl",
            "space_id": space_id,
            "file_id": file_id,
            "conversation_info": conversation_result,
            "upload": upload_result,
            "download": download_result,
        }
    return {
        "ok": True,
        "space_id": space_id,
        "file_id": file_id,
        "download_url": download_url,
        "conversation_info": conversation_result,
        "upload": upload_result,
        "download": download_result,
    }


def item_timestamp_ms(item: dict[str, Any]) -> int | None:
    for key in ("modifyTime", "createTime"):
        value = item.get(key)
        if value is None or value == "":
            continue
        try:
            return int(float(str(value)))
        except (TypeError, ValueError):
            continue
    return None


def cleanup_uploaded_dingtalk_images(
    dws_command: str,
    space_id: str,
    current_file_id: str,
    timeout_sec: int,
    dingtalk: dict[str, Any],
    *,
    force_current_delete: bool = False,
) -> dict[str, Any]:
    enabled = bool_value(dingtalk.get("cleanup_uploaded_images"), False)
    if not enabled:
        return {"ok": True, "enabled": False}

    prefix = str(dingtalk.get("uploaded_image_name_prefix") or "token-usage-").strip()
    retention_hours = max(float_value(dingtalk.get("uploaded_image_retention_hours"), 24.0), 0.0)
    max_cleanup = min(max(int_value(dingtalk.get("uploaded_image_max_cleanup"), 20), 0), 30)
    max_pages = min(max(int_value(dingtalk.get("uploaded_image_cleanup_max_pages"), 5), 1), 20)
    delete_current = force_current_delete or bool_value(dingtalk.get("delete_current_uploaded_image"), False)

    if not prefix:
        return {"ok": False, "enabled": True, "error": "missing uploaded_image_name_prefix"}
    if max_cleanup == 0:
        return {"ok": True, "enabled": True, "deleted": [], "skipped": [{"reason": "max_cleanup_is_zero"}]}

    now_ms = int(time.time() * 1000)
    cutoff_ms = now_ms - int(retention_hours * 3600 * 1000)
    deleted: list[dict[str, Any]] = []
    failed: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    scanned = 0
    pages = 0
    next_token = ""

    while pages < max_pages and len(deleted) + len(failed) < max_cleanup:
        command = [
            dws_command,
            "drive",
            "list",
            "--space-id",
            space_id,
            "--max",
            "50",
            "--order-by",
            "modifyTime",
            "--order",
            "desc",
            "--format",
            "json",
        ]
        if next_token:
            command.extend(["--next-token", next_token])
        list_result = run_dws_json(command, timeout_sec)
        pages += 1
        if not dws_success(list_result):
            return {
                "ok": False,
                "enabled": True,
                "error": "drive list failed",
                "deleted": deleted,
                "failed": failed,
                "skipped": skipped,
                "list": list_result,
            }

        payload = payload_result(list_result)
        items = payload.get("items")
        if not isinstance(items, list):
            items = []

        for item in items:
            if not isinstance(item, dict):
                continue
            scanned += 1
            name = str(item.get("name") or "")
            if not name.startswith(prefix) or not name.lower().endswith(".png"):
                continue

            file_id = str(item.get("fileId") or item.get("dentryId") or "").strip()
            timestamp_ms = item_timestamp_ms(item)
            age_hours = round((now_ms - timestamp_ms) / 3600000, 2) if timestamp_ms else None
            summary = {"name": name, "file_id": file_id, "age_hours": age_hours}

            if not file_id:
                skipped.append({**summary, "reason": "missing_file_id"})
                continue
            if current_file_id and file_id == current_file_id and not delete_current:
                skipped.append({**summary, "reason": "current_upload"})
                continue
            is_current_forced = bool(current_file_id and file_id == current_file_id and delete_current)
            if not is_current_forced and retention_hours > 0 and timestamp_ms and timestamp_ms > cutoff_ms:
                skipped.append({**summary, "reason": "within_retention"})
                continue

            delete_result = run_dws_json(
                [dws_command, "drive", "delete", "--node", file_id, "--yes", "--format", "json"],
                timeout_sec,
            )
            if dws_success(delete_result):
                deleted.append(summary)
            else:
                failed.append({**summary, "result": delete_result})
            if len(deleted) + len(failed) >= max_cleanup:
                break

        next_token = str(payload.get("nextToken") or "").strip()
        if not next_token:
            break

    return {
        "ok": not failed,
        "enabled": True,
        "space_id": space_id,
        "prefix": prefix,
        "retention_hours": retention_hours,
        "delete_current": delete_current,
        "scanned": scanned,
        "pages": pages,
        "deleted": deleted,
        "failed": failed,
        "skipped": skipped[:20],
        "skipped_count": len(skipped),
    }


def send_dingtalk_message(config: dict[str, Any], text: str, at: dict[str, list[str]], report: dict[str, Any]) -> dict[str, Any]:
    dingtalk = object_field(config, "dingtalk")
    mode = str(dingtalk.get("mode") or "current_user")
    dws_command = str(dingtalk.get("dws_command") or "dws")
    title = str(dingtalk.get("title") or "每日 Token 消耗提醒")
    timeout_sec = int(dingtalk.get("timeout_sec", 60))
    command = [dws_command, "chat", "message"]

    if mode == "current_user":
        group = env_or_config(dingtalk, "group_open_conversation_id", "DINGTALK_GROUP_OPEN_CONVERSATION_ID")
        if not group:
            raise TokenUsageError("missing dingtalk.group_open_conversation_id", code="config_error")
        command.extend(
            [
                "send",
                "--group",
                group,
                "--title",
                title,
                "--text",
                text,
            ]
        )
    elif mode == "bot":
        robot_code = env_or_config(dingtalk, "robot_code", "DINGTALK_ROBOT_CODE")
        group = env_or_config(dingtalk, "group_open_conversation_id", "DINGTALK_GROUP_OPEN_CONVERSATION_ID")
        if not robot_code or not group:
            raise TokenUsageError("missing dingtalk.robot_code or dingtalk.group_open_conversation_id", code="config_error")
        command.extend(["send-by-bot", "--robot-code", robot_code, "--group", group, "--title", title, "--text", text])
        if at["user_ids"]:
            command.extend(["--at-user-ids", ",".join(sorted(set(at["user_ids"])))])
        elif at["open_dingtalk_ids"]:
            command.extend(["--at-open-dingtalk-ids", ",".join(sorted(set(at["open_dingtalk_ids"])))])
    elif mode == "webhook":
        token = env_or_config(dingtalk, "webhook_token", "DINGTALK_WEBHOOK_TOKEN")
        if not token:
            raise TokenUsageError("missing dingtalk.webhook_token", code="config_error")
        command.extend(["send-by-webhook", "--token", token, "--title", title, "--text", text])
        if at["mobiles"]:
            command.extend(["--at-mobiles", ",".join(sorted(set(at["mobiles"])))])
        elif at["user_ids"]:
            command.extend(["--at-users", ",".join(sorted(set(at["user_ids"])))])
    else:
        raise TokenUsageError(f"unsupported dingtalk.mode: {mode}", code="config_error")

    command.extend(["--format", "json"])
    safe_command = ["<redacted>" if item in {env_or_config(dingtalk, "webhook_token", "DINGTALK_WEBHOOK_TOKEN")} else item for item in command]
    result = run_dws_json(command, timeout_sec)
    result["command"] = safe_command
    return result


def send_dingtalk_report(
    config: dict[str, Any],
    markdown: str,
    at: dict[str, list[str]],
    report: dict[str, Any],
    artifacts: dict[str, str],
) -> dict[str, Any]:
    dingtalk = object_field(config, "dingtalk")
    image_delivery = str(dingtalk.get("image_delivery") or "text").strip().lower()
    mode = str(dingtalk.get("mode") or "current_user")
    dws_command = str(dingtalk.get("dws_command") or "dws")
    timeout_sec = int(dingtalk.get("timeout_sec", 60))
    group = env_or_config(dingtalk, "group_open_conversation_id", "DINGTALK_GROUP_OPEN_CONVERSATION_ID")
    title = str(dingtalk.get("title") or "每日 AI 金额消耗提醒")
    image_path = Path(artifacts["image"])

    if image_delivery == "markdown_upload" and mode == "current_user" and group:
        image_result = upload_dingtalk_image(dws_command, image_path, group, timeout_sec)
        if image_result.get("ok"):
            image_markdown = f"![{title}]({image_result['download_url']})"
            send_result = send_dingtalk_message(config, image_markdown, {"open_dingtalk_ids": [], "user_ids": [], "mobiles": []}, report)
            send_result["delivery"] = "markdown_upload_image"
            send_result["image_upload"] = image_result
            if send_result["returncode"] == 0:
                send_result["image_cleanup"] = cleanup_uploaded_dingtalk_images(
                    dws_command,
                    str(image_result.get("space_id") or ""),
                    str(image_result.get("file_id") or ""),
                    timeout_sec,
                    dingtalk,
                )
                return send_result
            fallback_result = send_dingtalk_message(config, markdown, at, report)
            fallback_result["delivery"] = "text_fallback_after_image_markdown_send_failed"
            fallback_result["image_upload"] = image_result
            fallback_result["image_send"] = send_result
            if fallback_result["returncode"] == 0:
                fallback_result["image_cleanup"] = cleanup_uploaded_dingtalk_images(
                    dws_command,
                    str(image_result.get("space_id") or ""),
                    str(image_result.get("file_id") or ""),
                    timeout_sec,
                    dingtalk,
                    force_current_delete=True,
                )
            return fallback_result
        fallback_result = send_dingtalk_message(config, markdown, at, report)
        fallback_result["delivery"] = "text_fallback_after_image_upload_failed"
        fallback_result["image_upload"] = image_result
        return fallback_result

    send_result = send_dingtalk_message(config, markdown, at, report)
    send_result["delivery"] = "text"
    return send_result


def write_artifacts(config: dict[str, Any], report: dict[str, Any], markdown: str) -> dict[str, str]:
    outputs = object_field(config, "outputs") if isinstance(config.get("outputs"), dict) else {}
    output_dir = resolve_path(str(outputs.get("dir") or "outputs/token_usage_reports"))
    output_dir.mkdir(parents=True, exist_ok=True)
    report_date = report["report_date"]
    json_path = output_dir / f"token-usage-{report_date}.json"
    md_path = output_dir / f"token-usage-{report_date}.md"
    image_path = output_dir / f"token-usage-{report_date}.png"
    json_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    md_path.write_text(markdown, encoding="utf-8")
    render_report_image(report, config, image_path)
    return {"json": str(json_path), "markdown": str(md_path), "image": str(image_path)}


def run(args: argparse.Namespace) -> dict[str, Any]:
    load_local_env()
    config_path = resolve_path(args.config)
    config = read_json(config_path)
    if args.source_csv:
        config.setdefault("source", {})
        config["source"]["type"] = "csv"
        config["source"]["path"] = args.source_csv
    if args.source_xlsx:
        config.setdefault("source", {})
        config["source"]["type"] = "xlsx"
        config["source"]["path"] = args.source_xlsx

    window = parse_report_window(config, args.date)
    started = time.perf_counter()
    usage, identity_field = fetch_usage(config, window)
    source = object_field(config, "source")
    daily_usage: dict[str, UsageRecord] | None = None
    if str(source.get("type") or "") == "smartbi_task" and source.get("daily_zero_task"):
        daily_usage = fetch_smartbi_task_usage(source, window, task_override=str(source.get("daily_zero_task")))
    report = build_report(config, usage, identity_field, window, daily_usage=daily_usage)
    markdown, at = build_markdown(report, config)
    artifacts = write_artifacts(config, report, markdown)

    send_result: dict[str, Any] | None = None
    send_skipped_reason = None
    if args.send:
        if should_send(report, config):
            send_result = send_dingtalk_report(config, markdown, at, report, artifacts)
            if send_result["returncode"] != 0:
                raise TokenUsageError(f"DingTalk send failed: {send_result['stderr']}", code="dingtalk_error")
        else:
            send_skipped_reason = "send_when_all_ok=false and no low consumers"
    else:
        send_skipped_reason = "dry_run"

    return {
        "status": "ok",
        "mode": "send" if args.send else "dry_run",
        "config": str(config_path),
        "duration_sec": round(time.perf_counter() - started, 3),
        "report": {
            "report_date": report["report_date"],
            "counts": report["counts"],
            "total_tokens": report["total_tokens"],
            "total_amount": report.get("total_amount"),
            "low_consumers": report["low_consumers"],
            "daily_zero_consumers": report.get("daily_zero_consumers", []),
        },
        "artifacts": artifacts,
        "dingtalk": {
            "sent": send_result is not None,
            "skipped_reason": send_skipped_reason,
            "send_result": send_result,
        },
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Daily token usage DingTalk reporter")
    parser.add_argument("--config", default=str(DEFAULT_CONFIG), help="Path to JSON config.")
    parser.add_argument("--date", help="Report date in local timezone, YYYY-MM-DD. Defaults to yesterday.")
    parser.add_argument("--source-csv", help="Override source to read usage from this CSV file.")
    parser.add_argument("--source-xlsx", help="Override source to read usage from this Excel file.")
    parser.add_argument("--send", action="store_true", help="Actually send the DingTalk message. Default is dry-run.")
    parser.add_argument("--json", action="store_true", help="Print machine-readable JSON.")
    return parser


def main(argv: list[str] | None = None) -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    if hasattr(sys.stderr, "reconfigure"):
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    args = build_parser().parse_args(argv)
    try:
        result = run(args)
    except TokenUsageError as error:
        payload = {"status": "error", "error": {"code": error.code, "message": str(error)}}
        if args.json:
            print(json.dumps(payload, ensure_ascii=False, indent=2), file=sys.stderr)
        else:
            print(f"ERROR[{error.code}]: {error}", file=sys.stderr)
        return 2
    except Exception as error:  # pragma: no cover - defensive CLI boundary
        payload = {"status": "error", "error": {"code": "unexpected_error", "message": str(error)}}
        if args.json:
            print(json.dumps(payload, ensure_ascii=False, indent=2), file=sys.stderr)
        else:
            print(f"ERROR[unexpected_error]: {error}", file=sys.stderr)
        return 2

    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        report = result["report"]
        print(
            f"{report['report_date']}: total={format_number(int(report['total_tokens']))}, "
            f"low={report['counts']['low_consumers']}, mode={result['mode']}"
        )
        print(f"Markdown: {result['artifacts']['markdown']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
