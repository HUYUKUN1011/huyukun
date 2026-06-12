#!/usr/bin/env python3
"""SmartBI Data CLI: read-only SmartBI export CLI.

The CLI supports the original single-report MVP and a config-driven task runner.
Task configs keep report IDs, filter choices, and output paths outside code so
new BI reports can be added without changing this file.
"""

from __future__ import annotations

import argparse
import asyncio
import datetime as dt
import hashlib
import http.cookiejar
import importlib.util
import json
import mimetypes
import os
import re
import ssl
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
import zipfile
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


BASE_URL = "https://bi.61info.cn/smartbi/vision"
ROOT = Path(__file__).resolve().parents[1]
REPORT_ID = "I2c92808701977a217a21f809019785f3deca42cb"
REPORT_PATH = "分析报表/海外直播业务线/海外产运/外呼/益智外呼质量监控"
DEFAULT_CONFIG = Path("configs/smartbi_tasks.json")
DEFAULT_WRITEBACK_CONFIG = Path("configs/smartbi_writeback_tasks.json")
LOCAL_ENV_PATH = ROOT / ".env"
LOCAL_ENV_KEYS = {"SMARTBI_USERNAME", "SMARTBI_PASSWORD", "SMARTBI_BROWSER_CHANNEL"}
USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36"
)


def load_local_env(path: Path = LOCAL_ENV_PATH) -> None:
    if not path.exists():
        return
    try:
        lines = path.read_text(encoding="utf-8").splitlines()
    except OSError:
        return
    for raw_line in lines:
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if line.startswith("export "):
            line = line[len("export ") :].strip()
        if "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip().lstrip("\ufeff")
        if key not in LOCAL_ENV_KEYS or os.environ.get(key):
            continue
        value = value.strip()
        if len(value) >= 2 and value[0] == value[-1] and value[0] in {"'", '"'}:
            value = value[1:-1]
        os.environ[key] = value


class SmartbiError(RuntimeError):
    def __init__(self, message: str, code: str = "smartbi_error") -> None:
        super().__init__(message)
        self.code = code


class StepTimer:
    def __init__(self) -> None:
        self.started_at = time.perf_counter()
        self.steps: list[dict[str, Any]] = []

    def mark(self, name: str, started_at: float) -> None:
        self.steps.append({"name": name, "duration_sec": round(time.perf_counter() - started_at, 3)})

    def summary(self) -> dict[str, Any]:
        return {
            "total_sec": round(time.perf_counter() - self.started_at, 3),
            "steps": self.steps,
        }


class SmartbiClient:
    def __init__(self, base_url: str = BASE_URL) -> None:
        self.base_url = base_url.rstrip("/")
        self.cookie_jar = http.cookiejar.CookieJar()
        self.opener = urllib.request.build_opener(
            urllib.request.HTTPCookieProcessor(self.cookie_jar),
            urllib.request.HTTPSHandler(context=ssl.create_default_context()),
        )

    def request(
        self,
        path: str,
        data: dict[str, str] | None = None,
        referer: str | None = None,
    ) -> tuple[urllib.response.addinfourl, bytes]:
        url = path if path.startswith("http") else f"{self.base_url}/{path.lstrip('/')}"
        headers = {
            "User-Agent": USER_AGENT,
            "Referer": referer or f"{self.base_url}/index.jsp",
        }
        body = None
        if data is not None:
            body = urllib.parse.urlencode(data).encode("utf-8")
            headers["Content-Type"] = "application/x-www-form-urlencoded; charset=UTF-8"
            headers["Origin"] = "https://bi.61info.cn"
        req = urllib.request.Request(url, data=body, headers=headers)
        last_error: Exception | None = None
        for attempt in range(1, 4):
            try:
                with self.opener.open(req, timeout=120) as response:
                    return response, response.read()
            except (urllib.error.URLError, TimeoutError, ssl.SSLError, OSError) as error:
                last_error = error
                if attempt == 3:
                    break
                time.sleep(attempt)
        raise SmartbiError(f"Smartbi request failed after retries: {last_error}", code="network_error")

    def request_raw(
        self,
        path: str,
        body: bytes,
        headers: dict[str, str],
        referer: str | None = None,
        retries: int = 3,
    ) -> tuple[urllib.response.addinfourl, bytes]:
        url = path if path.startswith("http") else f"{self.base_url}/{path.lstrip('/')}"
        request_headers = {
            "User-Agent": USER_AGENT,
            "Referer": referer or f"{self.base_url}/index.jsp",
            "Origin": "https://bi.61info.cn",
            **headers,
        }
        req = urllib.request.Request(url, data=body, headers=request_headers)
        last_error: Exception | None = None
        for attempt in range(1, max(1, retries) + 1):
            try:
                with self.opener.open(req, timeout=120) as response:
                    return response, response.read()
            except (urllib.error.URLError, TimeoutError, ssl.SSLError, OSError) as error:
                last_error = error
                if attempt == max(1, retries):
                    break
                time.sleep(attempt)
        raise SmartbiError(f"Smartbi raw request failed after retries: {last_error}", code="network_error")

    def rmi(self, class_name: str, method_name: str, params: list[object]) -> dict[str, object]:
        _, body = self.request(
            "RMIServlet",
            {
                "className": class_name,
                "methodName": method_name,
                "params": json.dumps(params, ensure_ascii=False),
            },
        )
        payload = json.loads(body.decode("utf-8"))
        if payload.get("retCode") not in (0, "0"):
            raise SmartbiError(f"RMI failed: {class_name}.{method_name}: {payload.get('retCode')}", code="smartbi_rmi_error")
        return payload

    def login(self, username: str, password: str) -> None:
        self.request("index.jsp?time=1778858593300")
        payload = self.rmi("UserService", "clickLogin", [username, password])
        if payload.get("result") is not True:
            raise SmartbiError("Smartbi login failed", code="auth_error")

    def root_elements(self) -> list[dict[str, Any]]:
        payload = self.rmi("CatalogService", "getRootElements", [])
        result = payload.get("result")
        if not isinstance(result, list):
            raise SmartbiError("CatalogService.getRootElements returned no result list", code="catalog_error")
        return [element for element in result if isinstance(element, dict)]

    def child_elements(self, parent_id: str) -> list[dict[str, Any]]:
        payload = self.rmi("CatalogService", "getChildElements", [parent_id])
        result = payload.get("result")
        if not isinstance(result, list):
            raise SmartbiError("CatalogService.getChildElements returned no result list", code="catalog_error")
        return [element for element in result if isinstance(element, dict)]

    def open_report_context(self, report_id: str) -> dict[str, object]:
        report_url = f"openresource.jsp?isBrowse=true&showLeftTree=default&resid={report_id}"
        _, body = self.request(report_url, referer=f"{self.base_url}/index.jsp?time=1778858593300")
        html = body.decode("utf-8", "replace")
        raw_context = extract_js_object(html, "var spreadsheetReportContext =")
        return json.loads(raw_context)

    def export_spreadsheet_report(
        self,
        report_id: str,
        context: dict[str, object],
        params: list[dict[str, Any]] | None = None,
    ) -> tuple[str, bytes]:
        user_param_info = context.get("userParamInfo")
        if not isinstance(user_param_info, str):
            raise SmartbiError("Report context did not include userParamInfo")
        if params is None:
            params = json.loads(user_param_info or "[]")
        visible_sheets = context.get("visibleSheetNames") or []
        post_data = {
            "sheetIndex": str(context.get("activeSheetIndex", 0)),
            "resid": report_id,
            "clientId": required_str(context, "clientId"),
            "refreshType": "EXCEL2007",
            "paramsInfo": json.dumps(params, ensure_ascii=False),
            "pageId": "",
            "writeBackData": "",
            "exportSheetIndexes": ",".join(str(name) for name in visible_sheets),
            "exportFormula": "true",
            "tabControlsState": "",
        }
        response, body = self.request(
            "ssreportServlet",
            post_data,
            referer=f"{self.base_url}/openresource.jsp?isBrowse=true&showLeftTree=default&resid={report_id}",
        )
        content_type = response.headers.get("Content-Type", "")
        if "spreadsheetml.sheet" not in content_type and not body.startswith(b"PK\x03\x04"):
            preview = body[:300].decode("utf-8", "replace")
            raise SmartbiError(f"Export did not return an Excel file: {content_type}: {preview}", code="export_error")
        return default_filename(response.headers.get("Content-Disposition")) or f"{context.get('alias')}.xlsx", body


def load_config(path: Path) -> dict[str, Any]:
    if path.suffix.lower() in {".yaml", ".yml"}:
        raise SmartbiError("YAML config is not enabled in V1; use JSON to keep the CLI dependency-free", code="config_error")
    with path.open("r", encoding="utf-8") as handle:
        data = json.load(handle)
    if not isinstance(data, dict):
        raise SmartbiError("Config root must be a JSON object", code="config_error")
    tasks = data.get("tasks")
    if not isinstance(tasks, dict) or not tasks:
        raise SmartbiError("Config must include a non-empty 'tasks' object", code="config_error")
    return data


def get_task(config: dict[str, Any], task_name: str) -> dict[str, Any]:
    tasks = config.get("tasks")
    if not isinstance(tasks, dict):
        raise SmartbiError("Config missing tasks", code="config_error")
    task = tasks.get(task_name)
    if not isinstance(task, dict):
        available = ", ".join(sorted(str(name) for name in tasks))
        raise SmartbiError(f"Unknown task '{task_name}'. Available tasks: {available}", code="config_error")
    if task.get("enabled") is False:
        raise SmartbiError(f"Task '{task_name}' is disabled", code="config_error")
    return task


def task_report(task: dict[str, Any]) -> dict[str, Any]:
    report = task.get("report")
    if not isinstance(report, dict):
        raise SmartbiError("Task missing report object", code="config_error")
    if not isinstance(report.get("id"), str) or not report["id"]:
        raise SmartbiError("Task report.id is required", code="config_error")
    return report


def params_from_context(context: dict[str, object]) -> list[dict[str, Any]]:
    raw = context.get("userParamInfo")
    if not isinstance(raw, str):
        raise SmartbiError("Report context did not include userParamInfo", code="parameter_error")
    params = json.loads(raw or "[]")
    if not isinstance(params, list):
        raise SmartbiError("Report userParamInfo must be a parameter list", code="parameter_error")
    return merge_output_parameters(params, context)


def merge_output_parameters(params: list[dict[str, Any]], context: dict[str, object]) -> list[dict[str, Any]]:
    effective = [dict(param) for param in params]
    seen_ids = {str(param.get("id", "")) for param in effective}
    output_parameters = context.get("outputParameters") or []
    if not isinstance(output_parameters, list):
        return effective
    for parameter in output_parameters:
        if not isinstance(parameter, dict):
            continue
        parameter_id = str(parameter.get("id") or "")
        if not parameter_id or parameter_id in seen_ids:
            continue
        effective.append(
            {
                "id": parameter_id,
                "name": str(parameter.get("name") or ""),
                "alias": str(parameter.get("alias") or parameter.get("name") or ""),
                "value": "" if parameter.get("value") is None else str(parameter.get("value")),
                "displayValue": ""
                if parameter.get("displayValue") is None
                else str(parameter.get("displayValue")),
            }
        )
        seen_ids.add(parameter_id)
    return effective


def apply_param_overrides(params: list[dict[str, Any]], overrides: list[dict[str, Any]]) -> list[dict[str, Any]]:
    effective = [dict(param) for param in params]
    for override in overrides:
        if not isinstance(override, dict):
            raise SmartbiError("Each filter override must be an object")
        key = str(override.get("key") or "")
        if not key:
            raise SmartbiError("Filter override missing key")
        target = find_param(effective, key)
        value = override.get("value", "")
        display_value = override.get("displayValue", value)
        target["value"] = "" if value is None else str(value)
        target["displayValue"] = "" if display_value is None else str(display_value)
    return effective


def append_extra_params(params: list[dict[str, Any]], extra_params: list[dict[str, Any]]) -> list[dict[str, Any]]:
    effective = [dict(param) for param in params]
    for extra in extra_params:
        if not isinstance(extra, dict):
            raise SmartbiError("Each extra param must be an object")
        param_id = str(extra.get("id") or extra.get("key") or "")
        name = str(extra.get("name") or extra.get("key") or "")
        if not param_id or not name:
            raise SmartbiError("Each extra param requires id and name")
        value = extra.get("value", "")
        display_value = extra.get("displayValue", value)
        effective.append(
            {
                "id": param_id,
                "name": name,
                "alias": str(extra.get("alias") or name),
                "value": "" if value is None else str(value),
                "displayValue": "" if display_value is None else str(display_value),
            }
        )
    return effective


def find_param(params: list[dict[str, Any]], key: str) -> dict[str, Any]:
    for param in params:
        candidates = [
            str(param.get("id", "")),
            str(param.get("name", "")),
            str(param.get("alias", "")),
        ]
        if key in candidates:
            return param
    available = ", ".join(
        sorted({str(param.get("alias") or param.get("name") or param.get("id")) for param in params})
    )
    raise SmartbiError(f"Cannot find filter '{key}'. Available filters: {available}")


def apply_date_window(params: list[dict[str, Any]], date_window: str | None, today: dt.date | None = None) -> list[dict[str, Any]]:
    if not date_window:
        return params
    start, end = resolve_date_window(date_window, today=today)
    effective = [dict(param) for param in params]
    set_first_available_param_value(effective, ("start_date", "开始日期"), start.isoformat())
    set_first_available_param_value(effective, ("end_date", "结束日期"), end.isoformat())
    return effective


def resolve_date_window(date_window: str, today: dt.date | None = None) -> tuple[dt.date, dt.date]:
    today = today or dt.date.today()
    if date_window == "current_day":
        return today, today
    if date_window == "previous_day":
        yesterday = today - dt.timedelta(days=1)
        return yesterday, yesterday
    if date_window == "previous_week":
        this_monday = today - dt.timedelta(days=today.weekday())
        start = this_monday - dt.timedelta(days=7)
        end = this_monday - dt.timedelta(days=1)
        return start, end
    if date_window == "current_week_snapshot":
        start = today - dt.timedelta(days=today.weekday())
        return start, today
    if date_window == "current_month_snapshot":
        return today.replace(day=1), today
    if date_window == "previous_month":
        first_this_month = today.replace(day=1)
        end = first_this_month - dt.timedelta(days=1)
        start = end.replace(day=1)
        return start, end
    raise SmartbiError(f"Unsupported date_window: {date_window}")


def set_param_value(params: list[dict[str, Any]], key: str, value: str) -> None:
    param = find_param(params, key)
    param["value"] = value
    param["displayValue"] = value


def set_first_available_param_value(params: list[dict[str, Any]], keys: tuple[str, ...], value: str) -> None:
    for key in keys:
        try:
            set_param_value(params, key, value)
            return
        except SmartbiError:
            continue
    set_param_value(params, keys[0], value)


def task_filter_overrides(task: dict[str, Any]) -> list[dict[str, Any]]:
    filters = task.get("filters") or {}
    if not isinstance(filters, dict):
        raise SmartbiError("Task filters must be an object")
    overrides = filters.get("overrides") or []
    if not isinstance(overrides, list):
        raise SmartbiError("Task filters.overrides must be a list")
    return overrides


def task_extra_params(task: dict[str, Any]) -> list[dict[str, Any]]:
    filters = task.get("filters") or {}
    if not isinstance(filters, dict):
        raise SmartbiError("Task filters must be an object")
    extra_params = filters.get("extra_params") or []
    if not isinstance(extra_params, list):
        raise SmartbiError("Task filters.extra_params must be a list")
    return extra_params


def task_date_window(task: dict[str, Any]) -> str | None:
    filters = task.get("filters") or {}
    if not isinstance(filters, dict):
        raise SmartbiError("Task filters must be an object")
    value = filters.get("date_window")
    if value is None:
        return None
    if not isinstance(value, str):
        raise SmartbiError("Task filters.date_window must be a string")
    return value


def resolve_out_dir(template: str, task_name: str, run_date: dt.date, run_id: str, config_path: Path) -> Path:
    rendered = template.format(task=task_name, run_date=run_date.isoformat(), run_id=run_id)
    path = Path(rendered).expanduser()
    if path.is_absolute():
        return path
    return (ROOT / path).resolve()


def task_output_dir(task: dict[str, Any], task_name: str, run_date: dt.date, run_id: str, config_path: Path) -> Path:
    output = task.get("output") or {}
    if not isinstance(output, dict):
        raise SmartbiError("Task output must be an object")
    template = output.get("dir") or "outputs/bi_exports/{task}/{run_date}"
    if not isinstance(template, str):
        raise SmartbiError("Task output.dir must be a string")
    return resolve_out_dir(template, task_name, run_date, run_id, config_path)


def validate_xlsx(path: Path) -> dict[str, Any]:
    if path.stat().st_size <= 0:
        raise SmartbiError(f"Exported file is empty: {path}")
    if not zipfile.is_zipfile(path):
        raise SmartbiError(f"Exported file is not a valid xlsx zip: {path}")
    with zipfile.ZipFile(path) as workbook:
        bad_member = workbook.testzip()
        if bad_member:
            raise SmartbiError(f"XLSX zip validation failed at member: {bad_member}")
        sheet_names = extract_workbook_sheet_names(workbook)
    return {"bytes": path.stat().st_size, "sheets": sheet_names}


def extract_workbook_sheet_names(workbook: zipfile.ZipFile) -> list[str]:
    try:
        import xml.etree.ElementTree as ET

        xml = workbook.read("xl/workbook.xml")
        root = ET.fromstring(xml)
        namespace = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        return [sheet.attrib.get("name", "") for sheet in root.findall("main:sheets/main:sheet", namespace)]
    except Exception:
        return []


def load_export_rows(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=False)
    if "Sheet1" not in wb.sheetnames:
        raise SmartbiError(f"Exported workbook missing Sheet1: {path}")
    ws = wb["Sheet1"]

    headers = [ws.cell(2, col).value for col in range(1, ws.max_column + 1)]
    merged: dict[tuple[int, int], Any] = {}
    for merged_range in ws.merged_cells.ranges:
        value = ws.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged[(row, col)] = value

    rows: list[dict[str, Any]] = []
    for row in range(3, ws.max_row + 1):
        item: dict[str, Any] = {}
        has_data = False
        for col, header in enumerate(headers, start=1):
            if header in (None, "", 1.0):
                continue
            value = ws.cell(row, col).value
            if value is None:
                value = merged.get((row, col))
            if value not in (None, ""):
                has_data = True
            item[str(header)] = value
        if has_data:
            rows.append(item)
    return rows


def normalize_expected_values(value: object) -> list[str]:
    if value is None:
        return []
    if isinstance(value, str):
        return [part.strip() for part in value.split(",") if part.strip()]
    if isinstance(value, (list, tuple, set)):
        return [str(part).strip() for part in value if str(part).strip()]
    return [str(value).strip()] if str(value).strip() else []


def validate_task_assertions(task: dict[str, Any], export_path: Path) -> dict[str, Any] | None:
    assertions = task.get("assertions")
    if not assertions:
        return None
    if not isinstance(assertions, dict):
        raise SmartbiError("Task assertions must be an object when present")

    dimensions = assertions.get("dimensions") or {}
    if not isinstance(dimensions, dict):
        raise SmartbiError("Task assertions.dimensions must be an object when present")

    rows = load_export_rows(export_path)
    distinct: dict[str, list[str]] = {}
    errors: list[str] = []
    warnings: list[str] = []

    for dim, rule in dimensions.items():
        if not isinstance(dim, str) or not dim:
            errors.append("assertions.dimensions contains empty key")
            continue
        values = {str(row.get(dim)) for row in rows if row.get(dim) not in (None, "")}
        distinct[dim] = sorted(values)

        only: list[str] = []
        require: list[str] = []
        if isinstance(rule, dict):
            only = normalize_expected_values(rule.get("only"))
            require = normalize_expected_values(rule.get("require"))
        else:
            only = normalize_expected_values(rule)

        if only:
            unexpected = sorted(v for v in values if v not in set(only))
            if unexpected:
                errors.append(f"{dim}: unexpected values {unexpected} (only={only})")
        if require:
            missing = sorted(v for v in set(require) if v not in values)
            if missing:
                errors.append(f"{dim}: missing required values {missing} (require={require})")

        if dim not in rows[0] if rows else True:
            warnings.append(f"{dim}: header not found in export; assertion may be invalid")

    return {"ok": not errors, "errors": errors, "warnings": warnings, "distinct": distinct}


def run_task(
    task_name: str,
    task: dict[str, Any],
    config_path: Path,
    username: str | None,
    password: str | None,
    dry_run: bool,
    overwrite: bool,
    run_date: dt.date | None = None,
) -> dict[str, Any]:
    timer = StepTimer()
    report = task_report(task)
    run_date = run_date or dt.date.today()
    run_id = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    out_dir = task_output_dir(task, task_name, run_date, run_id, config_path)
    date_window = task_date_window(task)
    resolved_window = None
    if date_window:
        start_date, end_date = resolve_date_window(date_window, today=run_date)
        resolved_window = {"start_date": start_date.isoformat(), "end_date": end_date.isoformat()}
    plan = {
        "task": task_name,
        "report_id": report["id"],
        "report_path": report.get("path"),
        "date_window": date_window,
        "resolved_date_window": resolved_window,
        "extra_params": task_extra_params(task),
        "filter_overrides": task_filter_overrides(task),
        "out_dir": str(out_dir),
        "run_id": run_id,
    }
    if dry_run:
        return {"status": "dry_run", "plan": plan}
    if not username or not password:
        raise SmartbiError("SMARTBI_USERNAME/SMARTBI_PASSWORD or --username/--password is required")

    step_started = time.perf_counter()
    client = SmartbiClient()
    timer.mark("create_client", step_started)

    step_started = time.perf_counter()
    client.login(username, password)
    timer.mark("login", step_started)

    step_started = time.perf_counter()
    context = client.open_report_context(report["id"])
    timer.mark("open_report_context", step_started)

    step_started = time.perf_counter()
    params = params_from_context(context)
    params = append_extra_params(params, task_extra_params(task))
    params = apply_date_window(params, task_date_window(task), today=run_date)
    params = apply_param_overrides(params, task_filter_overrides(task))
    timer.mark("prepare_params", step_started)

    step_started = time.perf_counter()
    filename, body = client.export_spreadsheet_report(report["id"], context, params=params)
    timer.mark("export_spreadsheet_report", step_started)

    step_started = time.perf_counter()
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = unique_path(out_dir / filename, overwrite)
    out_path.write_bytes(body)
    timer.mark("write_output", step_started)

    step_started = time.perf_counter()
    validation = validate_xlsx(out_path)
    timer.mark("validate_xlsx", step_started)

    step_started = time.perf_counter()
    assertions_result = validate_task_assertions(task, out_path)
    timer.mark("validate_assertions", step_started)
    result = {
        **plan,
        "status": "exported",
        "output": str(out_path),
        "default_filename": filename,
        "bytes": validation["bytes"],
        "sheets": validation["sheets"],
        "assertions": assertions_result,
        "timings": timer.summary(),
    }
    if assertions_result and not assertions_result.get("ok", True):
        result["status"] = "error"
        write_run_log(out_dir, result)
        raise SmartbiError("Export assertions failed", code="assertion_error")
    write_run_log(out_dir, result)
    return result


def write_run_log(out_dir: Path, result: dict[str, Any]) -> None:
    log_path = out_dir / "run.json"
    log_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")


def load_writeback_task(config_path: Path, task_name: str) -> dict[str, Any]:
    with config_path.open("r", encoding="utf-8") as handle:
        config = json.load(handle)
    tasks = config.get("tasks")
    if not isinstance(tasks, dict):
        raise SmartbiError("Writeback config missing tasks object", code="config_error")
    task = tasks.get(task_name)
    if not isinstance(task, dict):
        available = ", ".join(sorted(str(name) for name in tasks))
        raise SmartbiError(f"Unknown writeback task '{task_name}'. Available tasks: {available}", code="config_error")
    if task.get("enabled") is False:
        raise SmartbiError(f"Writeback task '{task_name}' is disabled", code="config_error")
    return task


def writeback_target(task: dict[str, Any]) -> dict[str, Any]:
    target = task.get("target")
    if not isinstance(target, dict):
        raise SmartbiError("Writeback task missing target object", code="config_error")
    target_id = target.get("report_id")
    if not isinstance(target_id, str) or not target_id:
        raise SmartbiError("Writeback target.report_id is required", code="config_error")
    if target.get("type") != "DAQ_IMPORTCONFIG":
        raise SmartbiError("Writeback target.type must be DAQ_IMPORTCONFIG", code="config_error")
    return target


def response_header_dict(response: urllib.response.addinfourl) -> dict[str, str]:
    return {key: value for key, value in response.headers.items()}


def header_subset(headers: dict[str, str]) -> dict[str, str | None]:
    return {
        "content_type": headers.get("Content-Type"),
        "content_disposition": headers.get("Content-Disposition"),
        "content_length": headers.get("Content-Length"),
    }


def parse_import_config_page(html: str, target: dict[str, Any]) -> dict[str, Any]:
    target_id = str(target["report_id"])
    target_alias = str(target.get("alias") or "")
    script_res_id = extract_js_string_assignment(html, "resId")
    script_res_name = extract_js_string_assignment(html, "resName")
    checks = {
        "contains_target_id": target_id in html,
        "script_res_id_matches": script_res_id == target_id,
        "script_res_name_matches": script_res_name == target_alias,
        "has_excel_import_executor": "ExcelImportExecutorView" in html,
        "has_open_excel_import": "OPEN_EXCEL_IMPORT" in html,
        "has_daq_importconfig_type": "DAQ_IMPORTCONFIG" in html,
        "has_access_guard": "CUSTOM_DAQ" in html,
    }
    return {
        "script_res_id": script_res_id,
        "script_res_name": script_res_name,
        "checks": checks,
        "accepted": all(checks.values()),
    }


def extract_js_string_assignment(html: str, variable_name: str) -> str | None:
    match = re.search(rf"var\s+{re.escape(variable_name)}\s*=\s*'([^']*)'", html)
    if match:
        return match.group(1)
    return None


def build_future_upload_preview(target: dict[str, Any], input_filename: str | None = None) -> dict[str, Any]:
    filename = Path(input_filename).name if input_filename else "<input.xlsx>"
    return {
        "endpoint": "DataAcquisitionServlet",
        "method": "POST",
        "enctype": "multipart/form-data",
        "will_submit_in_http_probe": False,
        "fields": [
            {"name": "type", "value": "excelimport"},
            {
                "name": "id",
                "value": target["report_id"],
                "source": "task.target.report_id; rendered SmartBI form uses bofid=resId",
            },
            {"name": "parameterPanelBOId", "value": "", "source": "empty in current target probe"},
            {"name": "selectedRuleIds", "value": "", "source": "empty in current target probe"},
            {"name": "file", "filename": filename, "content": "[binary omitted]"},
        ],
        "blocked_until": [
            "owner provides edited test workbook",
            "owner provides original rollback workbook",
            "diff manifest passes",
            "owner confirms real upload window",
        ],
    }


def file_fingerprint(path: Path) -> dict[str, Any]:
    resolved = path.expanduser()
    digest = hashlib.sha256()
    with resolved.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return {
        "path": str(resolved),
        "name": resolved.name,
        "bytes": resolved.stat().st_size,
        "sha256": digest.hexdigest(),
    }


def build_data_acquisition_multipart_preview(
    *,
    target: dict[str, Any],
    candidate_file: Path,
) -> dict[str, Any]:
    """Build an auditable upload preview without encoding or submitting a body."""
    upload_fields = target.get("upload_fields") if isinstance(target.get("upload_fields"), dict) else {}
    file_info = file_fingerprint(candidate_file)
    return {
        "endpoint": "DataAcquisitionServlet",
        "method": "POST",
        "enctype": "multipart/form-data",
        "body_encoded": False,
        "request_sent": False,
        "upload_submitted": False,
        "fields": [
            {"name": "type", "value": "excelimport"},
            {
                "name": "id",
                "value": target["report_id"],
                "source": "task.target.report_id; DAQ_IMPORTCONFIG target id",
            },
            {
                "name": "parameterPanelBOId",
                "value": str(upload_fields.get("parameterPanelBOId") or ""),
                "source": "task.target.upload_fields.parameterPanelBOId or empty probe value",
            },
            {
                "name": "selectedRuleIds",
                "value": str(upload_fields.get("selectedRuleIds") or ""),
                "source": "task.target.upload_fields.selectedRuleIds or empty probe value",
            },
            {
                "name": "file",
                "filename": file_info["name"],
                "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "content": "[binary omitted]",
                "bytes": file_info["bytes"],
                "sha256": file_info["sha256"],
            },
        ],
    }


def summarize_upload_param(param: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": param.get("id"),
        "name": param.get("name"),
        "alias": param.get("alias"),
        "display": param.get("display"),
        "paramType": param.get("paramType"),
        "componentType": param.get("componentType"),
        "value_present": bool(param.get("value") or param.get("displayValue") or param.get("paramValue")),
    }


def initialize_data_acquisition_upload_context(client: SmartbiClient, target: dict[str, Any]) -> dict[str, Any]:
    target_id = str(target["report_id"])
    rules_payload = client.rmi("DataAcquisitionModule", "getImportConfigRules", [target_id])
    rules = rules_payload.get("result")
    if not isinstance(rules, list):
        raise SmartbiError("DataAcquisitionModule.getImportConfigRules returned no rule list", code="writeback_upload_context_failed")
    rule_items = [item for item in rules if isinstance(item, dict)]
    selected_rule_ids = ""
    select_ids_for_params = "all"
    if len(rule_items) > 1:
        selected_rule_ids = ",".join(str(item.get("id")) for item in rule_items if item.get("id") is not None)
        select_ids_for_params = selected_rule_ids or "all"

    params_payload = client.rmi("DataAcquisitionModule", "getAllParams", [target_id, select_ids_for_params])
    params_result = params_payload.get("result")
    if not isinstance(params_result, dict):
        raise SmartbiError("DataAcquisitionModule.getAllParams returned no parameter result", code="writeback_upload_context_failed")
    parameter_panel_bo_id = str(params_result.get("clientId") or "")
    parameter_panel_id = str(params_result.get("parameterPanelId") or "")
    params = params_result.get("params") if isinstance(params_result.get("params"), list) else []
    if params and not parameter_panel_bo_id:
        raise SmartbiError("DataAcquisitionModule.getAllParams returned params without clientId", code="writeback_upload_context_failed")

    return {
        "status": "ok",
        "mode": "data_acquisition_upload_context",
        "methods": [
            "DataAcquisitionModule.getImportConfigRules",
            "DataAcquisitionModule.getAllParams",
        ],
        "rules": [
            {
                "id": item.get("id"),
                "alias": item.get("alias"),
            }
            for item in rule_items
        ],
        "selectedRuleIds": selected_rule_ids,
        "selectIdsForParams": select_ids_for_params,
        "parameterPanelBOId": parameter_panel_bo_id,
        "parameterPanelId": parameter_panel_id,
        "params_count": len(params),
        "params": [summarize_upload_param(param) for param in params if isinstance(param, dict)],
    }


def target_with_upload_context(target: dict[str, Any], upload_context: dict[str, Any]) -> dict[str, Any]:
    effective = dict(target)
    configured_fields = target.get("upload_fields") if isinstance(target.get("upload_fields"), dict) else {}
    effective["upload_fields"] = {
        **configured_fields,
        "parameterPanelBOId": str(upload_context.get("parameterPanelBOId") or configured_fields.get("parameterPanelBOId") or ""),
        "selectedRuleIds": str(upload_context.get("selectedRuleIds") or configured_fields.get("selectedRuleIds") or ""),
    }
    return effective


def build_owner_approval_token_payload(
    *,
    task_name: str,
    run_id: str,
    candidate_sha256: str,
    rollback_sha256: str,
    expected_diff_sha256: str,
    upload_window_start: str,
    upload_window_end: str,
) -> dict[str, str]:
    return {
        "task": task_name,
        "run_id": run_id,
        "candidate_sha256": candidate_sha256,
        "rollback_sha256": rollback_sha256,
        "expected_diff_sha256": expected_diff_sha256,
        "upload_window_start": upload_window_start,
        "upload_window_end": upload_window_end,
    }


def build_owner_approval_token(payload: dict[str, str]) -> str:
    material = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(f"SMARTBI_WRITEBACK_OWNER_APPROVAL_V1:{material}".encode("utf-8")).hexdigest()


def build_owner_single_upload_token_payload(
    *,
    task_name: str,
    run_id: str,
    candidate_sha256: str,
    upload_window_start: str,
    upload_window_end: str,
) -> dict[str, str]:
    return {
        "mode": "SMARTBI_WRITEBACK_SINGLE_UPLOAD_V1",
        "task": task_name,
        "run_id": run_id,
        "candidate_sha256": candidate_sha256,
        "upload_window_start": upload_window_start,
        "upload_window_end": upload_window_end,
    }


def build_owner_single_upload_token(payload: dict[str, str]) -> str:
    material = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(f"SMARTBI_WRITEBACK_SINGLE_UPLOAD_V1:{material}".encode("utf-8")).hexdigest()


def parse_upload_window(value: str, name: str) -> dt.datetime:
    try:
        return dt.datetime.fromisoformat(value)
    except ValueError as error:
        raise SmartbiError(f"{name} must be ISO datetime, got {value!r}", code="upload_window_invalid") from error


def assert_upload_window(start: str, end: str, now: dt.datetime | None = None) -> dict[str, str]:
    start_at = parse_upload_window(start, "--upload-window-start")
    end_at = parse_upload_window(end, "--upload-window-end")
    current = now or dt.datetime.now()
    if end_at <= start_at:
        raise SmartbiError("upload window end must be after start", code="upload_window_invalid")
    if current < start_at or current > end_at:
        raise SmartbiError("current time is outside the approved upload window", code="upload_window_closed")
    return {
        "start": start_at.isoformat(timespec="seconds"),
        "end": end_at.isoformat(timespec="seconds"),
        "checked_at": current.isoformat(timespec="seconds"),
    }


def assert_sha256_lock(actual: str, expected: str | None, label: str) -> None:
    if not expected:
        raise SmartbiError(f"{label} sha256 lock is required", code="sha256_lock_required")
    if actual.lower() != expected.lower():
        raise SmartbiError(f"{label} sha256 mismatch", code="sha256_mismatch")


def build_multipart_form_data(
    *,
    target: dict[str, Any],
    candidate_file: Path,
    boundary: str,
) -> tuple[bytes, str]:
    upload_fields = target.get("upload_fields") if isinstance(target.get("upload_fields"), dict) else {}
    fields = [
        ("type", "excelimport"),
        ("id", str(target["report_id"])),
        ("parameterPanelBOId", str(upload_fields.get("parameterPanelBOId") or "")),
        ("selectedRuleIds", str(upload_fields.get("selectedRuleIds") or "")),
    ]
    chunks: list[bytes] = []
    for name, value in fields:
        chunks.append(f"--{boundary}\r\n".encode("utf-8"))
        chunks.append(f'Content-Disposition: form-data; name="{name}"\r\n\r\n'.encode("utf-8"))
        chunks.append(str(value).encode("utf-8"))
        chunks.append(b"\r\n")

    content_type = mimetypes.guess_type(candidate_file.name)[0] or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    chunks.append(f"--{boundary}\r\n".encode("utf-8"))
    chunks.append(
        (
            f'Content-Disposition: form-data; name="file"; filename="{candidate_file.name}"\r\n'
            f"Content-Type: {content_type}\r\n\r\n"
        ).encode("utf-8")
    )
    chunks.append(candidate_file.read_bytes())
    chunks.append(b"\r\n")
    chunks.append(f"--{boundary}--\r\n".encode("utf-8"))
    return b"".join(chunks), content_type


def submit_data_acquisition_multipart(
    *,
    client: SmartbiClient,
    target: dict[str, Any],
    candidate_file: Path,
    retries: int = 3,
) -> dict[str, Any]:
    boundary = f"----SmartBIWriteback{dt.datetime.now().strftime('%Y%m%d%H%M%S%f')}"
    body, file_content_type = build_multipart_form_data(target=target, candidate_file=candidate_file, boundary=boundary)
    response, response_body = client.request_raw(
        "DataAcquisitionServlet",
        body=body,
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
        referer=f"{client.base_url}/openimportconfig.jsp",
        retries=retries,
    )
    headers = response_header_dict(response)
    parsed = parse_data_acquisition_response(headers, response_body)
    if parsed["status"] == "ok":
        upload_result = "success_candidate_pending_verify"
    elif parsed["status"] == "error":
        upload_result = "failed"
    else:
        upload_result = "ambiguous_failed_stop"
    return {
        "status": "ok" if upload_result == "success_candidate_pending_verify" else "error",
        "upload_result": upload_result,
        "endpoint": "DataAcquisitionServlet",
        "method": "POST",
        "request_sent": True,
        "upload_submitted": True,
        "request_metadata": {
            "content_type": "multipart/form-data",
            "file_content_type": file_content_type,
            "file": file_fingerprint(candidate_file),
            "binary_logged": False,
        },
        "response": {
            "status": getattr(response, "status", None),
            "headers": header_subset(headers),
            "parsed": parsed,
        },
    }


def inspect_writeback_upload_workbook(path: Path) -> dict[str, Any]:
    def text(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, dt.datetime):
            return value.date().isoformat()
        if isinstance(value, dt.date):
            return value.isoformat()
        return str(value).strip()

    workbook = load_workbook(path.expanduser(), read_only=False, data_only=True)
    ws = workbook[workbook.sheetnames[0]]
    header_row = 3 if ws.max_row >= 3 else 1
    headers = [text(ws.cell(header_row, col).value) for col in range(1, ws.max_column + 1)]
    data_rows = 0
    for row_number in range(header_row + 1, ws.max_row + 1):
        values = [text(ws.cell(row_number, col).value) for col in range(1, ws.max_column + 1)]
        if any(values):
            data_rows += 1
    return {
        "path": str(path.expanduser()),
        "sheets": workbook.sheetnames,
        "active_sheet": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "assumed_header_row": header_row,
        "headers": headers,
        "data_rows": data_rows,
    }


def parse_decimal_value(value: Any) -> Decimal:
    if value is None:
        return Decimal("0")
    if isinstance(value, str) and not value.strip():
        return Decimal("0")
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return Decimal("0")


def workbook_cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    text = str(value).strip()
    if text.endswith(" 00:00:00"):
        return text[:10]
    return text


def find_probable_header_row(ws: Any, max_scan_rows: int = 20) -> tuple[int, list[str]]:
    best_row = 1
    best_headers: list[str] = []
    best_score = -1
    known_headers = {"日期", "平台", "投放平台", "投放账户", "广告位", "课包", "国家英文名称", "曝光", "点击", "消耗"}
    for row_number in range(1, min(ws.max_row, max_scan_rows) + 1):
        headers = [workbook_cell_text(ws.cell(row_number, col).value) for col in range(1, ws.max_column + 1)]
        non_empty = [header for header in headers if header]
        score = len(non_empty) + (len(set(non_empty) & known_headers) * 10)
        if score > best_score:
            best_row = row_number
            best_headers = headers
            best_score = score
    return best_row, best_headers


def load_workbook_rows(path: Path, header_row: int | None = None, start_col: int = 1) -> dict[str, Any]:
    workbook = load_workbook(path.expanduser(), read_only=False, data_only=True)
    ws = workbook[workbook.sheetnames[0]]
    if header_row is None:
        header_row, _ = find_probable_header_row(ws)
    headers = [workbook_cell_text(ws.cell(header_row, col).value) for col in range(start_col, ws.max_column + 1)]
    header_pairs = [
        (header, index + start_col)
        for index, header in enumerate(headers)
        if header
    ]
    rows = []
    for row_number in range(header_row + 1, ws.max_row + 1):
        row = {header: ws.cell(row_number, col).value for header, col in header_pairs}
        if all(workbook_cell_text(value) == "" for value in row.values()):
            continue
        row["_row_number"] = row_number
        rows.append(row)
    return {
        "path": str(path.expanduser()),
        "sheet": ws.title,
        "sheets": workbook.sheetnames,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "header_row": header_row,
        "headers": [header for header, _ in header_pairs],
        "rows": rows,
    }


def summarize_writeback_workbook(path: Path, task: dict[str, Any] | None = None) -> dict[str, Any]:
    loaded = load_workbook_rows(path)
    headers = loaded["headers"]
    rows = loaded["rows"]
    date_column = "日期" if "日期" in headers else None
    platform_column = "平台" if "平台" in headers else "投放平台" if "投放平台" in headers else None
    default_key_columns = ["日期", platform_column, "投放账户", "广告位", "课包", "国家英文名称"]
    key_columns = [str(column) for column in default_key_columns if column and column in headers]
    metric_columns = [column for column in ["曝光", "点击", "消耗"] if column in headers]
    dates = sorted({workbook_cell_text(row.get(date_column)) for row in rows if date_column and workbook_cell_text(row.get(date_column))})
    key_values = {
        "|".join(workbook_cell_text(row.get(column)) for column in key_columns)
        for row in rows
    } if key_columns else set()
    metric_sums = {
        column: format(sum((parse_decimal_value(row.get(column)) for row in rows), Decimal("0")), "f")
        for column in metric_columns
    }
    schema_errors: list[str] = []
    if task:
        schema = task.get("schema") or {}
        configured_header_sets = [
            [str(header) for header in schema.get("expected_headers", [])]
        ]
        input_cfg = task.get("input") if isinstance(task.get("input"), dict) else {}
        accepted_layouts = input_cfg.get("accepted_layouts") if isinstance(input_cfg.get("accepted_layouts"), list) else []
        for layout in accepted_layouts:
            if isinstance(layout, dict) and isinstance(layout.get("expected_headers"), list):
                configured_header_sets.append([str(header) for header in layout["expected_headers"]])
        configured_header_sets = [header_set for header_set in configured_header_sets if header_set]
        if configured_header_sets and not any(headers[: len(header_set)] == header_set for header_set in configured_header_sets):
            schema_errors.append("headers do not match any configured task schema")
    return {
        "status": "ok" if not schema_errors else "warning",
        "mode": "writeback_check_local",
        "boundary": {
            "smartbi_login": False,
            "data_acquisition_servlet_request_sent": False,
            "upload_submitted": False,
            "external_write": False,
        },
        "workbook": {
            key: loaded[key]
            for key in ["path", "sheet", "sheets", "max_row", "max_column", "header_row", "headers"]
        },
        "row_count": len(rows),
        "date_range": {
            "min": dates[0] if dates else None,
            "max": dates[-1] if dates else None,
            "values": dates[:50],
        },
        "key_columns": key_columns,
        "key_count": len(key_values),
        "metric_sums": metric_sums,
        "schema_errors": schema_errors,
    }


def canonical_writeback_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "日期": workbook_cell_text(row.get("日期")),
        "平台": workbook_cell_text(row.get("平台") if "平台" in row else row.get("投放平台")),
        "投放账户": workbook_cell_text(row.get("投放账户")),
        "广告位": workbook_cell_text(row.get("广告位")),
        "课包": workbook_cell_text(row.get("课包")),
        "国家英文名称": workbook_cell_text(row.get("国家英文名称")),
        "曝光": parse_decimal_value(row.get("曝光")),
        "点击": parse_decimal_value(row.get("点击")),
        "消耗": parse_decimal_value(row.get("消耗")),
    }


def aggregate_writeback_rows(rows: list[dict[str, Any]], key_columns: list[str]) -> dict[tuple[str, ...], dict[str, Any]]:
    aggregated: dict[tuple[str, ...], dict[str, Any]] = {}
    for row in rows:
        canonical = canonical_writeback_row(row)
        key = tuple(str(canonical[column]) for column in key_columns)
        item = aggregated.setdefault(key, {"曝光": Decimal("0"), "点击": Decimal("0"), "消耗": Decimal("0"), "count": 0})
        for column in ["曝光", "点击", "消耗"]:
            item[column] += canonical[column]
        item["count"] += 1
    return aggregated


def compare_writeback_candidate_subset(candidate_file: Path, export_file: Path) -> dict[str, Any]:
    candidate = load_workbook_rows(candidate_file)
    export = load_workbook_rows(export_file, start_col=2)
    candidate_rows = candidate["rows"]
    export_rows = export["rows"]
    key_columns = ["日期", "平台", "投放账户", "广告位", "课包", "国家英文名称"]
    candidate_map = aggregate_writeback_rows(candidate_rows, key_columns)
    export_map = aggregate_writeback_rows(export_rows, key_columns)
    missing = sorted(candidate_map.keys() - export_map.keys())
    value_mismatches = []
    for key in sorted(candidate_map.keys() & export_map.keys()):
        for column in ["曝光", "点击", "消耗"]:
            if candidate_map[key][column] != export_map[key][column]:
                value_mismatches.append(
                    {
                        "key": key,
                        "column": column,
                        "candidate": str(candidate_map[key][column]),
                        "export": str(export_map[key][column]),
                    }
                )
    return {
        "status": "ok" if not missing and not value_mismatches else "mismatch",
        "mode": "post_verify_subset_compare",
        "candidate": {
            "path": str(candidate_file.expanduser()),
            "row_count": len(candidate_rows),
            "key_count": len(candidate_map),
        },
        "post_verify_export": {
            "path": str(export_file.expanduser()),
            "row_count": len(export_rows),
            "key_count": len(export_map),
        },
        "compare": {
            "strategy": "candidate subset",
            "key_columns": key_columns,
            "candidate_keys_missing_in_export_count": len(missing),
            "value_mismatch_count": len(value_mismatches),
            "missing_sample": missing[:10],
            "value_mismatch_sample": value_mismatches[:10],
        },
    }


def load_json_file(path: Path, code: str) -> dict[str, Any]:
    try:
        with path.expanduser().open("r", encoding="utf-8") as handle:
            data = json.load(handle)
    except FileNotFoundError as error:
        raise SmartbiError(f"JSON file not found: {path}", code=code) from error
    except json.JSONDecodeError as error:
        raise SmartbiError(f"Invalid JSON file {path}: {error}", code=code) from error
    if not isinstance(data, dict):
        raise SmartbiError(f"JSON file must contain an object: {path}", code=code)
    return data


def normalize_diff_for_match(diff: dict[str, Any]) -> dict[str, Any]:
    return {
        "status": diff.get("status"),
        "task": diff.get("task"),
        "key_columns": diff.get("key_columns"),
        "changed_row_count": diff.get("changed_row_count"),
        "changed_rows": diff.get("changed_rows"),
        "errors": diff.get("errors") or [],
    }


def verify_expected_diff(expected_diff: dict[str, Any], current_diff: dict[str, Any]) -> dict[str, Any]:
    if expected_diff.get("status") != "ok":
        raise SmartbiError("expected-diff must have status=ok", code="expected_diff_invalid")
    expected_normalized = normalize_diff_for_match(expected_diff)
    current_normalized = normalize_diff_for_match(current_diff)
    if expected_normalized != current_normalized:
        raise SmartbiError(
            "expected-diff does not match current rollback/candidate workbook diff",
            code="expected_diff_mismatch",
        )
    return {
        "status": "ok",
        "matched_fields": list(expected_normalized.keys()),
        "changed_row_count": current_diff.get("changed_row_count"),
        "expected_diff_status": expected_diff.get("status"),
    }


def extract_balanced_json_object(text: str, start_index: int) -> tuple[str | None, int | None]:
    brace = text.find("{", start_index)
    if brace < 0:
        return None, None
    depth = 0
    quote: str | None = None
    escaped = False
    for index in range(brace, len(text)):
        char = text[index]
        if quote:
            if escaped:
                escaped = False
            elif char == "\\":
                escaped = True
            elif char == quote:
                quote = None
            continue
        if char in {"'", '"'}:
            quote = char
        elif char == "{":
            depth += 1
        elif char == "}":
            depth -= 1
            if depth == 0:
                return text[brace : index + 1], index + 1
    return None, None


def extract_data_acquisition_callback(text: str) -> dict[str, Any]:
    marker = "doImportFormSubmitCallback"
    start = text.find(marker)
    if start < 0:
        return {
            "found": False,
            "json": None,
            "json_text": None,
            "error": None,
        }
    search_at = start
    last_error = None
    while True:
        brace = text.find("{", search_at)
        if brace < 0:
            break
        json_text, end = extract_balanced_json_object(text, brace)
        if not json_text or end is None:
            break
        try:
            payload = json.loads(json_text)
        except json.JSONDecodeError as error:
            last_error = f"callback_json_decode_error: {error}"
            search_at = brace + 1
            continue
        if any(key in payload for key in ("success", "result", "errorMessage", "sheetResults", "retCode")):
            return {
                "found": True,
                "json": payload,
                "json_text": json_text,
                "error": None,
                "end_index": end,
            }
        search_at = brace + 1
    if last_error:
        return {
            "found": True,
            "json": None,
            "json_text": None,
            "error": last_error,
        }
    if "{" in text[start:]:
        return {
            "found": True,
            "json": None,
            "json_text": None,
            "error": "callback_json_truncated_or_missing",
        }
    return {
        "found": True,
        "json": None,
        "json_text": None,
        "error": "callback_json_missing",
    }


def summarize_sheet_results(sheet_results: Any) -> dict[str, Any]:
    if not isinstance(sheet_results, list):
        return {
            "present": False,
            "sheet_count": 0,
            "failed_sheet_count": 0,
            "row_error_count": None,
            "column_error_count": None,
        }
    failed_sheet_count = 0
    row_error_count = 0
    column_error_count = 0
    for item in sheet_results:
        if not isinstance(item, dict):
            continue
        if item.get("success") is False or item.get("result") is False:
            failed_sheet_count += 1
        for key, value in item.items():
            lowered_key = str(key).lower()
            if "row" in lowered_key and isinstance(value, list):
                row_error_count += len(value)
            if "column" in lowered_key and isinstance(value, list):
                column_error_count += len(value)
    return {
        "present": True,
        "sheet_count": len(sheet_results),
        "failed_sheet_count": failed_sheet_count,
        "row_error_count": row_error_count,
        "column_error_count": column_error_count,
    }


def parse_data_acquisition_response(headers: dict[str, str], body: bytes) -> dict[str, Any]:
    content_type = headers.get("Content-Type") or headers.get("content-type") or ""
    disposition = headers.get("Content-Disposition") or headers.get("content-disposition") or ""
    text = body[:5000].decode("utf-8", "replace") if body else ""
    lowered = text.lower()
    json_payload = None
    try:
        json_payload = json.loads(text) if text.strip() else None
    except json.JSONDecodeError:
        json_payload = None
    callback = extract_data_acquisition_callback(text)
    callback_json = callback.get("json") if callback.get("found") else None
    callback_json_text = callback.get("json_text") if callback.get("found") else None
    effective_payload = callback_json if isinstance(callback_json, dict) else json_payload
    sheet_results = effective_payload.get("sheetResults") if isinstance(effective_payload, dict) else None
    sheet_results_summary = summarize_sheet_results(sheet_results)
    error_message = effective_payload.get("errorMessage") if isinstance(effective_payload, dict) else None
    exception_download = "下载异常数据" in text or "error data" in lowered or "exception data" in lowered
    looks_like_file = body.startswith(b"PK\x03\x04") or "attachment" in disposition.lower()
    success_markers = [
        "导入成功",
        "上传成功",
        "操作成功",
        "执行成功",
        "import success",
        "upload success",
        "successfully",
    ]
    failure_markers = [
        "导入失败",
        "上传失败",
        "操作失败",
        "执行失败",
        "异常",
        "error",
        "exception",
        "import failed",
        "upload failed",
    ]
    success = any(marker in text or marker in lowered for marker in success_markers)
    failure = any(marker in text or marker in lowered for marker in failure_markers)
    if isinstance(effective_payload, dict):
        if effective_payload.get("success") is True or effective_payload.get("result") is True:
            success = True
            failure = False
        if effective_payload.get("success") is False or effective_payload.get("result") is False:
            failure = True
        ret_code = effective_payload.get("retCode")
        if ret_code in (0, "0"):
            success = True
            failure = False
        elif ret_code not in (None, ""):
            failure = True
        if isinstance(error_message, str) and error_message.strip():
            failure = True
    if success and not failure:
        exception_download = False
    if not body:
        status = "empty"
    elif exception_download or failure:
        status = "error"
    elif success:
        status = "ok"
    elif looks_like_file:
        status = "file_response"
    else:
        status = "unknown"
    return {
        "status": status,
        "content_type": content_type or None,
        "content_disposition": disposition or None,
        "bytes": len(body),
        "has_exception_data_download": exception_download,
        "looks_like_file_response": looks_like_file,
        "json_detected": isinstance(json_payload, dict),
        "callback_detected": bool(callback.get("found")),
        "callback_json_extracted": isinstance(callback_json, dict),
        "callback_json": callback_json,
        "callback_json_text": callback_json_text,
        "callback_error": callback.get("error"),
        "errorMessage": error_message,
        "sheetResults": sheet_results,
        "sheetResults_summary": sheet_results_summary,
        "sanitized_callback_text": text if callback.get("found") else None,
        "preview": text[:5000],
    }


def response_parser_fixture_results() -> dict[str, Any]:
    return {
        "success_html": parse_data_acquisition_response(
            {"Content-Type": "text/html; charset=UTF-8"},
            "<html><body>导入成功</body></html>".encode("utf-8"),
        ),
        "failure_html": parse_data_acquisition_response(
            {"Content-Type": "text/html; charset=UTF-8"},
            "<html><body>导入失败，可下载异常数据</body></html>".encode("utf-8"),
        ),
        "upload_success_html": parse_data_acquisition_response(
            {"Content-Type": "text/html; charset=UTF-8"},
            "<script>alert('上传成功');</script>".encode("utf-8"),
        ),
        "success_json": parse_data_acquisition_response(
            {"Content-Type": "application/json; charset=UTF-8"},
            b'{"success": true}',
        ),
        "callback_failure_html": parse_data_acquisition_response(
            {"Content-Type": "text/html; charset=UTF-8"},
            (
                "<script>parent.doImportFormSubmitCallback(true,'excelimport','target',"
                '{"success":false,"errorMessage":"bad data","sheetResults":[{"sheetIndex":0,"success":false}]});</script>'
            ).encode("utf-8"),
        ),
        "callback_success_html": parse_data_acquisition_response(
            {"Content-Type": "text/html; charset=UTF-8"},
            (
                "<script>if(parent.doImportFormSubmitCallback){"
                "parent.doImportFormSubmitCallback(true,'excelimport','target',"
                '{"success":true,"errorMessage":"","sheetResults":[{"sheetIndex":0,"sheetName":"sheet1",'
                '"success":true,"ruleName":"导入规则1","successCount":36,"totalCount":36}]});}</script>'
            ).encode("utf-8"),
        ),
        "callback_nested_if_failure_html": parse_data_acquisition_response(
            {"Content-Type": "text/html; charset=UTF-8"},
            (
                "<script>if(parent.doImportFormSubmitCallback){"
                "parent.doImportFormSubmitCallback(true,'excelimport','target',"
                '{"success":false,"errorMessage":"","sheetResults":[{"sheetIndex":0,"sheetName":"sheet1",'
                '"success":false,"errorMessage":"BindingParam","ruleName":"导入规则1",'
                '"successCount":0,"totalCount":0}]});}</script>'
            ).encode("utf-8"),
        ),
        "empty_response": parse_data_acquisition_response({}, b""),
    }


def build_post_verify_plan(task: dict[str, Any]) -> dict[str, Any]:
    post_verify = task.get("post_verify") if isinstance(task.get("post_verify"), dict) else {}
    report = post_verify.get("report") if isinstance(post_verify.get("report"), dict) else {}
    parameters = post_verify.get("parameters") if isinstance(post_verify.get("parameters"), dict) else {}
    compare_columns = post_verify.get("compare_columns") if isinstance(post_verify.get("compare_columns"), list) else []
    key_columns = post_verify.get("key_columns") if isinstance(post_verify.get("key_columns"), list) else []
    report_id = str(report.get("report_id") or "")
    report_path = str(report.get("path") or "")
    return {
        "status": "planned" if report.get("report_id") else "missing_report",
        "report": report,
        "parameters": parameters,
        "key_columns": key_columns,
        "compare_columns": compare_columns,
        "future_inspect_report_command": [
            "python3",
            "scripts/smartbi_cli.py",
            "inspect-report",
            "--report-id",
            report_id or "<post_verify.report.report_id>",
            "--report-path",
            report_path or "<post_verify.report.path>",
            "--json",
        ],
        "future_export_command_shape": [
            "python3",
            "scripts/smartbi_cli.py",
            "run",
            "--task",
            "<future_post_verify_export_task>",
            "--json",
        ],
        "future_compare_strategy": {
            "source": "future exported post_verify.report workbook",
            "match_keys": key_columns,
            "compare_columns": compare_columns,
            "expected_values": "accepted current_diff.changed_rows from shadow execution plan",
        },
        "next_paths": [
            "inspect-report for report metadata refresh",
            "future export of post_verify.report after real upload",
            "compare exported report rows against accepted diff keys",
        ],
        "executed": False,
        "post_verify_executed": False,
    }


def build_writeback_readiness_packet(
    *,
    task_name: str,
    shadow_plan: dict[str, Any],
    post_verify_plan: dict[str, Any],
    source_artifacts: dict[str, str],
) -> dict[str, Any]:
    boundary = shadow_plan.get("boundary") if isinstance(shadow_plan.get("boundary"), dict) else {}
    multipart = shadow_plan.get("multipart_request") if isinstance(shadow_plan.get("multipart_request"), dict) else {}
    expected_diff = shadow_plan.get("expected_diff") if isinstance(shadow_plan.get("expected_diff"), dict) else {}
    current_diff = shadow_plan.get("current_diff") if isinstance(shadow_plan.get("current_diff"), dict) else {}
    http_probe = shadow_plan.get("http_probe") if isinstance(shadow_plan.get("http_probe"), dict) else {}
    upload_submitted = bool(shadow_plan.get("upload_submitted") or multipart.get("upload_submitted"))
    request_sent = bool(boundary.get("request_sent") or multipart.get("request_sent"))
    owner_approval_present = False
    prerequisites_ok = (
        shadow_plan.get("status") == "ok"
        and expected_diff.get("status") == "ok"
        and current_diff.get("status") == "ok"
        and http_probe.get("status") == "ok"
        and post_verify_plan.get("status") == "planned"
        and not upload_submitted
        and not request_sent
    )
    readiness_decision = "not_ready_for_real_upload_owner_approval_required"
    if prerequisites_ok:
        readiness_decision = "shadow_ready_not_upload_ready"
    return {
        "status": "ok" if prerequisites_ok else "blocked",
        "mode": "writeback_real_upload_readiness_packet_no_upload",
        "task": task_name,
        "readiness_decision": readiness_decision,
        "boundary": {
            "owner_approval_present": owner_approval_present,
            "upload_approval_present": False,
            "execute_enablement_present": False,
            "real_upload_still_blocked": True,
            "writeback_execute_blocked_expected": True,
            "request_sent": request_sent,
            "data_acquisition_servlet_request_sent": request_sent,
            "upload_submitted": upload_submitted,
            "external_write": False,
        },
        "source_artifacts": source_artifacts,
        "summary": {
            "candidate_dry_run": "ok" if shadow_plan.get("candidate_workbook") else "missing",
            "rollback_dry_run": "ok" if shadow_plan.get("rollback_workbook") else "missing",
            "expected_diff": expected_diff.get("status"),
            "shadow_execution_plan": shadow_plan.get("status"),
            "http_servlet_probe": http_probe.get("status"),
            "post_verify_dry_run_plan": post_verify_plan.get("status"),
        },
        "target": shadow_plan.get("target"),
        "candidate_workbook": shadow_plan.get("candidate_workbook"),
        "rollback_workbook": shadow_plan.get("rollback_workbook"),
        "expected_diff": expected_diff,
        "multipart_request": multipart,
        "post_verify_dry_run_plan": post_verify_plan,
        "rollback_rehearsal_command_shape": [
            "python3",
            "scripts/smartbi_cli.py",
            "writeback",
            "--task",
            task_name,
            "--shadow-execute",
            "--candidate-file",
            "<rollback workbook path>",
            "--rollback-file",
            "<edited workbook path>",
            "--expected-diff",
            "<rollback diff manifest>",
            "--json",
        ],
        "owner_approval_wording_placeholder": (
            "Owner must explicitly approve a real upload window and immediate rollback "
            "window in this thread before any execute implementation is enabled."
        ),
        "real_upload_window_checklist": [
            "owner is present and monitoring SmartBI",
            "candidate and rollback files are immutable and fingerprinted",
            "expected diff is reviewed and accepted",
            "post-verify report route is available",
            "rollback shadow rehearsal is prepared",
            "execute hard block remains until a separate implementation goal",
        ],
        "rollback_window_checklist": [
            "rollback workbook is the untouched original",
            "rollback command shape is reviewed before upload",
            "post-rollback compare plan is ready",
            "owner can verify aggregate values after rollback",
        ],
        "stop_conditions": [
            "owner approval absent",
            "upload window absent",
            "rollback window absent",
            "expected diff not status=ok",
            "candidate or rollback fingerprint changed",
            "target http-probe fails",
            "post-verify plan missing",
            "any request_sent/upload_submitted flag becomes true during shadow stage",
        ],
        "residual_risks": [
            "DataAcquisitionServlet real response remains unobserved because upload is blocked",
            "parameterPanelBOId and selectedRuleIds are currently empty based on probe evidence",
            "post-verify compare is planned but not executed before a real upload",
            "real rollback effectiveness cannot be proven without a controlled upload window",
        ],
        "not_owner_approval": True,
        "not_upload_approval": True,
        "not_execute_enablement": True,
    }


def run_writeback_shadow_execute(
    *,
    config_path: Path,
    task_name: str,
    username: str | None,
    password: str | None,
    candidate_file: Path | None,
    rollback_file: Path | None,
    expected_diff_file: Path | None,
    max_changed_rows: int,
    run_dir: Path,
) -> dict[str, Any]:
    if candidate_file is None:
        raise SmartbiError("writeback --shadow-execute requires --candidate-file", code="config_error")
    if rollback_file is None:
        raise SmartbiError("writeback --shadow-execute requires --rollback-file", code="config_error")
    if expected_diff_file is None:
        raise SmartbiError("writeback --shadow-execute requires --expected-diff", code="config_error")

    from diff_smartbi_writeback_workbooks import build_diff
    from validate_smartbi_writeback_input import get_task as get_writeback_task
    from validate_smartbi_writeback_input import load_config as load_writeback_config
    from validate_smartbi_writeback_input import validate_workbook

    config = load_writeback_config(config_path)
    task = get_writeback_task(config, task_name)
    target = writeback_target(task)
    candidate = candidate_file.expanduser()
    rollback = rollback_file.expanduser()
    expected_diff_path = expected_diff_file.expanduser()

    candidate_dry_run = validate_workbook(candidate, task_name, task)
    if candidate_dry_run.get("status") != "ok":
        raise SmartbiError("candidate workbook dry-run failed", code="writeback_shadow_preflight_failed")
    rollback_dry_run = validate_workbook(rollback, task_name, task)
    if rollback_dry_run.get("status") != "ok":
        raise SmartbiError("rollback workbook dry-run failed", code="writeback_shadow_preflight_failed")

    current_diff = build_diff(
        original=rollback,
        candidate=candidate,
        task_name=task_name,
        task=task,
        max_changed_rows=max_changed_rows,
    )
    if current_diff.get("status") != "ok":
        raise SmartbiError("current rollback/candidate diff failed", code="writeback_shadow_preflight_failed")
    expected_diff = load_json_file(expected_diff_path, code="expected_diff_invalid")
    expected_diff_check = verify_expected_diff(expected_diff, current_diff)

    http_probe = run_writeback_http_probe(
        config_path=config_path,
        task_name=task_name,
        username=username,
        password=password,
        input_filename=str(candidate),
    )
    if http_probe.get("status") != "ok":
        raise SmartbiError("target http-probe failed", code="writeback_shadow_preflight_failed")

    plan = {
        "status": "ok",
        "mode": "writeback_shadow_execute_no_upload",
        "boundary": {
            "smartbi_login": True,
            "browser_used": False,
            "dry_run_candidate": True,
            "dry_run_rollback": True,
            "diff_verified": True,
            "target_http_probe": True,
            "request_built": True,
            "request_sent": False,
            "upload_submitted": False,
            "external_write": False,
        },
        "task": task_name,
        "target": target,
        "candidate_workbook": file_fingerprint(candidate),
        "rollback_workbook": file_fingerprint(rollback),
        "expected_diff": {
            "path": str(expected_diff_path),
            "status": expected_diff.get("status"),
            "check": expected_diff_check,
        },
        "current_diff": current_diff,
        "multipart_request": build_data_acquisition_multipart_preview(target=target, candidate_file=candidate),
        "http_probe": {
            "status": http_probe.get("status"),
            "import_config": http_probe.get("import_config"),
            "template_download": http_probe.get("template_download"),
            "upload_submitted": False,
        },
        "response_parser_design": {
            "implemented": True,
            "fixtures": response_parser_fixture_results(),
            "real_response_parsed": False,
        },
        "post_verify_plan": build_post_verify_plan(task),
        "stop_conditions": [
            "candidate dry-run fails",
            "rollback dry-run fails",
            "current diff fails",
            "expected-diff missing, not status=ok, or mismatched",
            "target http-probe fails",
            "owner approval for real upload is absent",
        ],
        "upload_submitted": False,
    }
    plan_path = run_dir / "writeback_shadow_execution_plan.json"
    plan_path.write_text(json.dumps(plan, ensure_ascii=False, indent=2), encoding="utf-8")
    plan["artifact"] = str(plan_path)
    return plan


def has_guarded_execute_args(args: argparse.Namespace) -> bool:
    guarded_fields = [
        "candidate_file",
        "rollback_file",
        "expected_diff",
        "owner_approval_token",
        "upload_window_start",
        "upload_window_end",
        "candidate_sha256",
        "rollback_sha256",
        "expected_diff_sha256",
    ]
    return any(bool(getattr(args, field, None)) for field in guarded_fields)


def run_writeback_guarded_execute(
    *,
    args: argparse.Namespace,
    config_path: Path,
    run_id: str,
    run_dir: Path,
) -> dict[str, Any]:
    if not args.confirm_writeback:
        raise SmartbiError("writeback --execute requires --confirm-writeback", code="writeback_execute_blocked")
    if not args.candidate_file:
        raise SmartbiError("writeback guarded execute requires --candidate-file", code="config_error")
    if not args.rollback_file:
        raise SmartbiError("writeback guarded execute requires --rollback-file", code="config_error")
    if not args.expected_diff:
        raise SmartbiError("writeback guarded execute requires --expected-diff", code="config_error")
    if not args.owner_approval_token:
        raise SmartbiError("writeback guarded execute requires --owner-approval-token", code="owner_approval_required")
    if not args.upload_window_start or not args.upload_window_end:
        raise SmartbiError("writeback guarded execute requires upload window start/end", code="upload_window_required")

    from diff_smartbi_writeback_workbooks import build_diff
    from validate_smartbi_writeback_input import get_task as get_writeback_task
    from validate_smartbi_writeback_input import load_config as load_writeback_config
    from validate_smartbi_writeback_input import validate_workbook

    config = load_writeback_config(config_path)
    task = get_writeback_task(config, args.task)
    target = writeback_target(task)
    candidate = Path(args.candidate_file).expanduser()
    rollback = Path(args.rollback_file).expanduser()
    expected_diff_path = Path(args.expected_diff).expanduser()

    candidate_info = file_fingerprint(candidate)
    rollback_info = file_fingerprint(rollback)
    expected_diff_info = file_fingerprint(expected_diff_path)
    assert_sha256_lock(str(candidate_info["sha256"]), args.candidate_sha256, "candidate")
    assert_sha256_lock(str(rollback_info["sha256"]), args.rollback_sha256, "rollback")
    assert_sha256_lock(str(expected_diff_info["sha256"]), args.expected_diff_sha256, "expected-diff")
    upload_window = assert_upload_window(args.upload_window_start, args.upload_window_end)

    token_payload = build_owner_approval_token_payload(
        task_name=args.task,
        run_id=run_id,
        candidate_sha256=str(candidate_info["sha256"]),
        rollback_sha256=str(rollback_info["sha256"]),
        expected_diff_sha256=str(expected_diff_info["sha256"]),
        upload_window_start=args.upload_window_start,
        upload_window_end=args.upload_window_end,
    )
    expected_token = build_owner_approval_token(token_payload)
    if args.owner_approval_token != expected_token:
        raise SmartbiError("owner approval token mismatch", code="owner_approval_token_mismatch")

    candidate_dry_run = validate_workbook(candidate, args.task, task)
    if candidate_dry_run.get("status") != "ok":
        raise SmartbiError("candidate workbook dry-run failed", code="writeback_execute_preflight_failed")
    rollback_dry_run = validate_workbook(rollback, args.task, task)
    if rollback_dry_run.get("status") != "ok":
        raise SmartbiError("rollback workbook dry-run failed", code="writeback_execute_preflight_failed")

    current_diff = build_diff(
        original=rollback,
        candidate=candidate,
        task_name=args.task,
        task=task,
        max_changed_rows=args.max_changed_rows,
    )
    if current_diff.get("status") != "ok":
        raise SmartbiError("current rollback/candidate diff failed", code="writeback_execute_preflight_failed")
    expected_diff = load_json_file(expected_diff_path, code="expected_diff_invalid")
    expected_diff_check = verify_expected_diff(expected_diff, current_diff)
    post_verify_plan = build_post_verify_plan(task)
    if post_verify_plan.get("status") != "planned":
        raise SmartbiError("post-verify dry-run plan is missing report metadata", code="writeback_execute_preflight_failed")

    client = login_client(args.username, args.password)
    target_id = str(target["report_id"])
    import_response, import_body = client.request(
        "openimportconfig.jsp",
        {
            "isBrowse": "true",
            "showLeftTree": "default",
            "resid": target_id,
        },
    )
    import_probe = parse_import_config_page(import_body.decode("utf-8", "replace"), target)
    if not import_probe.get("accepted"):
        raise SmartbiError("target import config validation failed", code="writeback_execute_preflight_failed")

    preflight = {
        "status": "ok",
        "mode": "writeback_guarded_execute_preflight",
        "task": args.task,
        "run_id": run_id,
        "target": target,
        "candidate_workbook": candidate_info,
        "rollback_workbook": rollback_info,
        "expected_diff": {
            "file": expected_diff_info,
            "status": expected_diff.get("status"),
            "check": expected_diff_check,
        },
        "upload_window": upload_window,
        "owner_approval": {
            "token_payload": token_payload,
            "token_verified": True,
        },
        "import_config": {
            "endpoint": "openimportconfig.jsp",
            "status": getattr(import_response, "status", None),
            **import_probe,
        },
        "multipart_request": build_data_acquisition_multipart_preview(target=target, candidate_file=candidate),
        "post_verify_plan": post_verify_plan,
        "boundary": {
            "request_sent": False,
            "data_acquisition_servlet_request_sent": False,
            "upload_submitted": False,
            "external_write": False,
        },
    }
    preflight_path = run_dir / "guarded_execute_preflight.json"
    preflight_path.write_text(json.dumps(preflight, ensure_ascii=False, indent=2), encoding="utf-8")

    upload_response = submit_data_acquisition_multipart(client=client, target=target, candidate_file=candidate)
    response_path = run_dir / "guarded_execute_response.json"
    response_path.write_text(json.dumps(upload_response, ensure_ascii=False, indent=2), encoding="utf-8")

    result = {
        "status": "ok" if upload_response.get("status") == "ok" else "error",
        "mode": "writeback_guarded_execute",
        "task": args.task,
        "run_id": run_id,
        "target": target,
        "upload_result": upload_response.get("upload_result"),
        "boundary": {
            "smartbi_login": True,
            "browser_used": False,
            "request_sent": True,
            "data_acquisition_servlet_request_sent": True,
            "upload_submitted": True,
            "external_write": True,
            "auto_retry": False,
            "auto_rollback": False,
        },
        "artifacts": {
            "directory": str(run_dir),
            "preflight": str(preflight_path),
            "response": str(response_path),
        },
        "post_verify_command": post_verify_plan.get("future_inspect_report_command"),
        "rollback_command_shape": [
            "python3",
            "scripts/smartbi_cli.py",
            "writeback",
            "--task",
            args.task,
            "--execute",
            "--confirm-writeback",
            "--candidate-file",
            str(rollback),
            "--rollback-file",
            str(candidate),
            "--expected-diff",
            "<rollback_expected_diff.json>",
            "--owner-approval-token",
            "<rollback-run-id-scoped-owner-token>",
            "--run-id",
            "<rollback-run-id>",
            "--json",
        ],
    }
    summary_path = run_dir / "writeback_run.json"
    summary_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    result["artifacts"]["summary"] = str(summary_path)
    return result


def run_writeback_single_upload(
    *,
    args: argparse.Namespace,
    config_path: Path,
    run_id: str,
    run_dir: Path,
) -> dict[str, Any]:
    if not args.confirm_writeback:
        raise SmartbiError("writeback --single-upload requires --confirm-writeback", code="writeback_execute_blocked")
    if not args.candidate_file:
        raise SmartbiError("writeback --single-upload requires --candidate-file", code="config_error")
    if not args.candidate_sha256:
        raise SmartbiError("writeback --single-upload requires --candidate-sha256", code="sha256_lock_required")
    if not args.owner_approval_token:
        raise SmartbiError("writeback --single-upload requires --owner-approval-token", code="owner_approval_required")
    if not args.upload_window_start or not args.upload_window_end:
        raise SmartbiError("writeback --single-upload requires upload window start/end", code="upload_window_required")

    task = load_writeback_task(config_path, args.task)
    target = writeback_target(task)
    candidate = Path(args.candidate_file).expanduser()
    candidate_info = file_fingerprint(candidate)
    assert_sha256_lock(str(candidate_info["sha256"]), args.candidate_sha256, "candidate")
    upload_window = assert_upload_window(args.upload_window_start, args.upload_window_end)
    token_payload = build_owner_single_upload_token_payload(
        task_name=args.task,
        run_id=run_id,
        candidate_sha256=str(candidate_info["sha256"]),
        upload_window_start=args.upload_window_start,
        upload_window_end=args.upload_window_end,
    )
    expected_token = build_owner_single_upload_token(token_payload)
    if args.owner_approval_token != expected_token:
        raise SmartbiError("owner approval token mismatch", code="owner_approval_token_mismatch")

    workbook_inspection = inspect_writeback_upload_workbook(candidate)
    post_verify_plan = build_post_verify_plan(task)
    if post_verify_plan.get("status") != "planned":
        raise SmartbiError("post-verify dry-run plan is missing report metadata", code="writeback_execute_preflight_failed")

    client = login_client(args.username, args.password)
    target_id = str(target["report_id"])
    import_response, import_body = client.request(
        "openimportconfig.jsp",
        {
            "isBrowse": "true",
            "showLeftTree": "default",
            "resid": target_id,
        },
    )
    import_probe = parse_import_config_page(import_body.decode("utf-8", "replace"), target)
    if not import_probe.get("accepted"):
        raise SmartbiError("target import config validation failed", code="writeback_execute_preflight_failed")
    upload_context = initialize_data_acquisition_upload_context(client, target)
    effective_target = target_with_upload_context(target, upload_context)

    preflight = {
        "status": "ok",
        "mode": "writeback_single_upload_preflight",
        "task": args.task,
        "run_id": run_id,
        "target": target,
        "candidate_workbook": candidate_info,
        "workbook_inspection": workbook_inspection,
        "upload_window": upload_window,
        "owner_approval": {
            "token_payload": token_payload,
            "token_verified": True,
        },
        "import_config": {
            "endpoint": "openimportconfig.jsp",
            "status": getattr(import_response, "status", None),
            **import_probe,
        },
        "upload_context": upload_context,
        "multipart_request": build_data_acquisition_multipart_preview(target=effective_target, candidate_file=candidate),
        "post_verify_plan": post_verify_plan,
        "boundary": {
            "smartbi_login": True,
            "browser_used": False,
            "data_acquisition_servlet_request_sent": False,
            "upload_submitted": False,
            "single_upload_request": True,
            "auto_retry": False,
            "auto_rollback": False,
        },
    }
    preflight_path = run_dir / "single_upload_preflight.json"
    preflight_path.write_text(json.dumps(preflight, ensure_ascii=False, indent=2), encoding="utf-8")

    upload_response = submit_data_acquisition_multipart(
        client=client,
        target=effective_target,
        candidate_file=candidate,
        retries=1,
    )
    response_path = run_dir / "single_upload_response.json"
    response_path.write_text(json.dumps(upload_response, ensure_ascii=False, indent=2), encoding="utf-8")

    result = {
        "status": "ok" if upload_response.get("status") == "ok" else "error",
        "mode": "writeback_single_upload",
        "task": args.task,
        "run_id": run_id,
        "target": target,
        "upload_result": upload_response.get("upload_result"),
        "callback": upload_response.get("response", {}).get("parsed", {}),
        "boundary": {
            "smartbi_login": True,
            "browser_used": False,
            "data_acquisition_servlet_request_sent": True,
            "upload_submitted": True,
            "single_upload_request": True,
            "auto_retry": False,
            "auto_rollback": False,
        },
        "artifacts": {
            "directory": str(run_dir),
            "preflight": str(preflight_path),
            "response": str(response_path),
        },
        "post_verify_command": post_verify_plan.get("future_inspect_report_command"),
    }
    summary_path = run_dir / "writeback_single_upload_run.json"
    summary_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    result["artifacts"]["summary"] = str(summary_path)
    return result


def run_writeback_http_probe(
    *,
    config_path: Path,
    task_name: str,
    username: str | None,
    password: str | None,
    input_filename: str | None = None,
) -> dict[str, Any]:
    task = load_writeback_task(config_path, task_name)
    target = writeback_target(task)
    client = login_client(username, password)
    target_id = str(target["report_id"])

    import_response, import_body = client.request(
        "openimportconfig.jsp",
        {
            "isBrowse": "true",
            "showLeftTree": "default",
            "resid": target_id,
        },
    )
    import_html = import_body.decode("utf-8", "replace")
    import_probe = parse_import_config_page(import_html, target)

    template_response, template_body = client.request(
        "ExcelTemplateDownloadServlet",
        {
            "resId": target_id,
            "paramsInfo": "",
            "downloadAllData": "",
        },
    )
    template_headers = response_header_dict(template_response)
    template_filename = default_filename(template_headers.get("Content-Disposition"))
    template_validation = {
        "status": getattr(template_response, "status", None),
        **header_subset(template_headers),
        "default_filename": template_filename,
        "bytes": len(template_body),
        "is_xlsx_zip": template_body.startswith(b"PK\x03\x04"),
    }

    accepted = bool(import_probe["accepted"]) and bool(template_validation["is_xlsx_zip"])
    return {
        "status": "ok" if accepted else "error",
        "mode": "http_servlet_writeback_probe_no_upload",
        "boundary": {
            "smartbi_login": True,
            "browser_used": False,
            "template_downloaded": True,
            "file_selected": False,
            "upload_submitted": False,
            "external_write": False,
        },
        "task": task_name,
        "target": target,
        "import_config": {
            "endpoint": "openimportconfig.jsp",
            "method": "POST",
            "status": getattr(import_response, "status", None),
            "content_type": import_response.headers.get("Content-Type"),
            **import_probe,
        },
        "template_download": template_validation,
        "future_upload_request": build_future_upload_preview(target, input_filename=input_filename),
        "errors": [] if accepted else ["HTTP servlet probe did not pass all checks"],
    }


def run_writeback_check(args: argparse.Namespace) -> dict[str, Any]:
    task = None
    if args.task:
        try:
            task = load_writeback_task(Path(args.config).expanduser(), args.task)
        except SmartbiError:
            task = None
    result = summarize_writeback_workbook(Path(args.file).expanduser(), task=task)
    result["task"] = args.task
    result["operator"] = args.operator
    if args.out:
        out_path = Path(args.out).expanduser()
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
        result["artifact"] = str(out_path)
    return result


def summarize_data_acquisition_response_artifact(response: dict[str, Any]) -> dict[str, Any]:
    parsed = response.get("response", {}).get("parsed") if isinstance(response.get("response"), dict) else response.get("parsed")
    parsed = parsed if isinstance(parsed, dict) else {}
    callback_json = parsed.get("callback_json") if isinstance(parsed.get("callback_json"), dict) else {}
    sheet_results = parsed.get("sheetResults") if isinstance(parsed.get("sheetResults"), list) else callback_json.get("sheetResults")
    sheet_results = sheet_results if isinstance(sheet_results, list) else []
    first_sheet = sheet_results[0] if sheet_results and isinstance(sheet_results[0], dict) else {}
    return {
        "status": "success" if callback_json.get("success") is True or parsed.get("status") == "ok" else "failed",
        "upload_result": response.get("upload_result"),
        "http_status": response.get("response", {}).get("status") if isinstance(response.get("response"), dict) else None,
        "callback_detected": parsed.get("callback_detected"),
        "callback_json_extracted": parsed.get("callback_json_extracted"),
        "success": callback_json.get("success"),
        "errorMessage": parsed.get("errorMessage") or callback_json.get("errorMessage") or first_sheet.get("errorMessage"),
        "sheetResults": sheet_results,
        "sheetResults_summary": parsed.get("sheetResults_summary") or summarize_sheet_results(sheet_results),
        "has_exception_data_download": parsed.get("has_exception_data_download"),
    }


def run_writeback_diagnose(args: argparse.Namespace) -> dict[str, Any]:
    run_dir = Path(args.run_dir).expanduser()
    candidates = [
        "single_upload_response_reparsed.json",
        "single_upload_response.json",
        "guarded_execute_response.json",
        "cli_owner_upload_response.json",
    ]
    response_path = next((run_dir / name for name in candidates if (run_dir / name).exists()), None)
    if response_path is None:
        raise SmartbiError(f"No known upload response artifact found in {run_dir}", code="diagnose_artifact_missing")
    response = load_json_file(response_path, code="diagnose_artifact_invalid")
    summary = summarize_data_acquisition_response_artifact(response)
    result = {
        "status": "ok",
        "mode": "writeback_diagnose",
        "run_dir": str(run_dir),
        "response_artifact": str(response_path),
        "diagnosis": summary,
        "boundary": {
            "smartbi_login": False,
            "data_acquisition_servlet_request_sent": False,
            "upload_submitted": False,
            "external_write": False,
        },
    }
    if args.out:
        out_path = Path(args.out).expanduser()
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
        result["artifact"] = str(out_path)
    return result


def extract_report_id_from_url(url: str) -> str | None:
    parsed = urllib.parse.urlparse(url)
    query = urllib.parse.parse_qs(parsed.query)
    for key in ("resid", "resId", "id"):
        values = query.get(key)
        if values:
            return values[0]
    return None


def infer_writeback_task_config_draft(
    *,
    task_name: str,
    target: dict[str, Any],
    sample_summary: dict[str, Any],
    post_verify_path: str | None,
    post_verify_url: str | None,
) -> dict[str, Any]:
    headers = sample_summary["workbook"]["headers"]
    platform_column = "平台" if "平台" in headers else "投放平台" if "投放平台" in headers else None
    key_columns = [column for column in ["日期", platform_column, "投放账户", "广告位", "课包", "国家英文名称"] if column]
    compare_columns = [column for column in ["曝光", "点击", "消耗"] if column in headers]
    return {
        "tasks": {
            task_name: {
                "enabled": False,
                "kind": "SMARTBI_EXCEL_WRITEBACK",
                "target": target,
                "post_verify": {
                    "report": {
                        "path": post_verify_path or post_verify_url or "<post_verify_report_path_or_url>",
                        "report_id": extract_report_id_from_url(post_verify_url or "") or "<post_verify_report_id>",
                        "type": "SPREADSHEET_REPORT",
                    },
                    "parameters": {},
                    "key_columns": key_columns,
                    "compare_columns": compare_columns,
                },
                "input": {
                    "allowed_dirs": ["/Users/takuya/Desktop", "outputs/smartbi_writeback_inputs"],
                    "filename_pattern": f"^{re.escape(Path(sample_summary['workbook']['path']).name)}$",
                    "workbook_type": "xlsx",
                    "sheet_name": sample_summary["workbook"]["sheet"],
                    "header_row": sample_summary["workbook"]["header_row"],
                    "data_start_row": int(sample_summary["workbook"]["header_row"]) + 1,
                    "max_rows": 5000,
                },
                "schema": {
                    "key_columns": key_columns,
                    "expected_headers": headers,
                    "required_non_empty": key_columns,
                    "date_columns": ["日期"] if "日期" in headers else [],
                    "numeric_columns": compare_columns,
                    "business_summary_columns": [column for column in ["日期", platform_column, "课包"] if column],
                },
            }
        }
    }


def run_writeback_onboard(args: argparse.Namespace) -> dict[str, Any]:
    run_id = args.run_id or dt.datetime.now().strftime("onboard-%Y%m%d-%H%M%S")
    run_dir = Path(args.out_root).expanduser() / args.task / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    target_id = extract_report_id_from_url(args.target_url or "") if args.target_url else None
    target_path = args.target_path
    client = None if args.offline else login_client(args.username, args.password)
    if client and not target_id and target_path:
        target_element, _, _ = resolve_catalog_path(client, target_path)
        target_id = required_element_id(target_element)
    if not target_id:
        if not args.offline:
            raise SmartbiError("writeback-onboard requires --target-url with resid or --target-path", code="config_error")
        target_id = "<target_report_id>"
    target = {
        "alias": Path(target_path).name if target_path else args.task,
        "path": target_path or args.target_url or "",
        "report_id": target_id,
        "type": "DAQ_IMPORTCONFIG",
    }
    sample_summary = summarize_writeback_workbook(Path(args.sample_file).expanduser(), task=None)
    import_probe: dict[str, Any] = {"accepted": False, "status": "not_probed_offline"}
    upload_context: dict[str, Any] = {
        "status": "not_probed_offline",
        "methods": [
            "DataAcquisitionModule.getImportConfigRules",
            "DataAcquisitionModule.getAllParams",
        ],
        "parameterPanelBOId": "",
        "selectedRuleIds": "",
    }
    template_info: dict[str, Any] = {
        "path": None,
        "bytes": 0,
        "is_xlsx_zip": False,
        "status": "not_downloaded_offline",
    }
    if client:
        import_response, import_body = client.request(
            "openimportconfig.jsp",
            {
                "isBrowse": "true",
                "showLeftTree": "default",
                "resid": target_id,
            },
        )
        import_probe = {
            "status": getattr(import_response, "status", None),
            **parse_import_config_page(import_body.decode("utf-8", "replace"), target),
        }
        upload_context = initialize_data_acquisition_upload_context(client, target)
        template_response, template_body = client.request(
            "ExcelTemplateDownloadServlet",
            {
                "resId": target_id,
                "paramsInfo": "",
                "downloadAllData": "",
            },
        )
        template_path = run_dir / (default_filename(template_response.headers.get("Content-Disposition")) or "writeback_template.xlsx")
        template_path.write_bytes(template_body)
        template_info = {
            "path": str(template_path),
            "bytes": len(template_body),
            "is_xlsx_zip": template_body.startswith(b"PK\x03\x04"),
            "status": "downloaded",
        }
    draft = infer_writeback_task_config_draft(
        task_name=args.task,
        target=target,
        sample_summary=sample_summary,
        post_verify_path=args.post_verify_path,
        post_verify_url=args.post_verify_url,
    )
    draft_path = run_dir / "writeback_task_config_draft.json"
    draft_path.write_text(json.dumps(draft, ensure_ascii=False, indent=2), encoding="utf-8")
    shadow_upload_plan = build_data_acquisition_multipart_preview(
        target=target_with_upload_context(target, upload_context),
        candidate_file=Path(args.sample_file).expanduser(),
    )
    shadow_upload_plan_path = run_dir / "shadow_upload_plan.json"
    shadow_upload_plan_path.write_text(json.dumps(shadow_upload_plan, ensure_ascii=False, indent=2), encoding="utf-8")
    configured = bool(import_probe.get("accepted")) and bool(template_info.get("is_xlsx_zip"))
    report = {
        "status": "configured" if configured else "discovered",
        "mode": "writeback_onboard_no_upload",
        "task": args.task,
        "onboarding_status": "configured" if configured else "discovered",
        "approved_for_developer_use": False,
        "target": target,
        "import_config": import_probe,
        "template": template_info,
        "sample_inspect": sample_summary,
        "upload_context_probe": upload_context,
        "task_config_draft": str(draft_path),
        "shadow_upload_plan": str(shadow_upload_plan_path),
        "post_verify_plan": draft["tasks"][args.task]["post_verify"],
        "next_statuses": ["discovered", "configured", "validated", "approved_for_developer_use"],
        "boundary": {
            "smartbi_login": bool(client),
            "template_downloaded": bool(client),
            "data_acquisition_servlet_request_sent": False,
            "upload_submitted": False,
            "external_write": False,
        },
    }
    report_path = run_dir / "writeback_onboarding_report.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    report["artifact"] = str(report_path)
    return report


def run_writeback_upload(args: argparse.Namespace) -> dict[str, Any]:
    if not args.confirm:
        raise SmartbiError("writeback-upload requires --confirm", code="writeback_execute_blocked")
    run_id = args.run_id or dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    candidate = Path(args.file).expanduser()
    candidate_info = file_fingerprint(candidate)
    start = (dt.datetime.now() - dt.timedelta(minutes=5)).isoformat(timespec="seconds")
    end = (dt.datetime.now() + dt.timedelta(hours=1)).isoformat(timespec="seconds")
    payload = build_owner_single_upload_token_payload(
        task_name=args.task,
        run_id=run_id,
        candidate_sha256=str(candidate_info["sha256"]),
        upload_window_start=start,
        upload_window_end=end,
    )
    single_args = argparse.Namespace(
        **vars(args),
        candidate_file=str(candidate),
        candidate_sha256=str(candidate_info["sha256"]),
        owner_approval_token=build_owner_single_upload_token(payload),
        upload_window_start=start,
        upload_window_end=end,
        confirm_writeback=True,
    )
    run_dir = Path(args.out_root).expanduser() / args.task / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    upload_result = run_writeback_single_upload(
        args=single_args,
        config_path=Path(args.config).expanduser(),
        run_id=run_id,
        run_dir=run_dir,
    )
    upload_result["operator"] = args.operator
    if upload_result.get("status") != "ok":
        return upload_result
    task = load_writeback_task(Path(args.config).expanduser(), args.task)
    post_verify = task.get("post_verify") if isinstance(task.get("post_verify"), dict) else {}
    report = post_verify.get("report") if isinstance(post_verify.get("report"), dict) else {}
    report_id = report.get("report_id")
    if isinstance(report_id, str) and report_id:
        post_verify_parameters = post_verify.get("parameters")
        post_verify_parameters = post_verify_parameters if isinstance(post_verify_parameters, dict) else {}
        export_task = {
            "enabled": True,
            "report": {
                "id": report_id,
                "path": report.get("path"),
            },
            "filters": {
                "overrides": [
                    {"key": key, "value": value, "displayValue": value}
                    for key, value in post_verify_parameters.items()
                ]
            },
            "output": {"dir": str(run_dir / "post_verify_export")},
        }
        post_verify_export = run_task(
            f"{args.task}_post_verify",
            export_task,
            Path(args.config).expanduser(),
            args.username,
            args.password,
            dry_run=False,
            overwrite=True,
        )
        upload_result["post_verify_export"] = post_verify_export
        export_path = post_verify_export.get("output")
        if isinstance(export_path, str):
            compare = compare_writeback_candidate_subset(candidate, Path(export_path))
            compare_path = run_dir / "post_verify_compare_summary.json"
            compare_path.write_text(json.dumps(compare, ensure_ascii=False, indent=2), encoding="utf-8")
            upload_result["post_verify_compare"] = compare
            upload_result.setdefault("artifacts", {})["post_verify_compare"] = str(compare_path)
            if compare.get("status") != "ok":
                upload_result["status"] = "error"
                upload_result["error"] = {
                    "code": "post_verify_mismatch",
                    "message": "post-verify candidate subset compare failed",
                }
    else:
        upload_result["post_verify_export"] = {
            "status": "skipped",
            "reason": "post_verify.report.report_id is not configured",
        }
    return upload_result


def run_writeback_command(args: argparse.Namespace) -> dict[str, Any]:
    config_path = Path(args.config).expanduser()
    run_id = args.run_id or dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    out_root = Path(args.out_root).expanduser()
    run_dir = out_root / args.task / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    if args.single_upload:
        return run_writeback_single_upload(args=args, config_path=config_path, run_id=run_id, run_dir=run_dir)
    if args.execute:
        if not has_guarded_execute_args(args):
            raise SmartbiError(
                "Real SmartBI writeback is intentionally blocked in this CLI slice; "
                "guarded execute requires owner approval token, SHA locks, upload window, "
                "candidate file, rollback file, and expected diff.",
                code="writeback_execute_blocked",
            )
        return run_writeback_guarded_execute(args=args, config_path=config_path, run_id=run_id, run_dir=run_dir)
    if not args.dry_run and not args.probe and not args.diff and not args.http_probe and not args.shadow_execute:
        raise SmartbiError("writeback requires --dry-run, --probe, --http-probe, --diff, --shadow-execute, and/or --single-upload", code="config_error")

    result: dict[str, Any] = {
        "status": "ok",
        "mode": "writeback_no_upload",
        "task": args.task,
        "run_id": run_id,
        "boundary": {
            "dry_run_allowed": bool(args.dry_run),
            "probe_allowed": bool(args.probe),
            "http_probe_allowed": bool(args.http_probe),
            "diff_allowed": bool(args.diff),
            "shadow_execute_allowed": bool(args.shadow_execute),
            "execute_allowed": False,
            "file_selected": False,
            "upload_submitted": False,
            "external_write": False,
        },
        "artifacts": {"directory": str(run_dir)},
        "steps": {},
    }

    if args.dry_run:
        if not args.file:
            raise SmartbiError("writeback --dry-run requires --file", code="config_error")
        from validate_smartbi_writeback_input import get_task as get_writeback_task
        from validate_smartbi_writeback_input import load_config as load_writeback_config
        from validate_smartbi_writeback_input import validate_workbook

        config = load_writeback_config(config_path)
        task = get_writeback_task(config, args.task)
        dry_run_result = validate_workbook(Path(args.file).expanduser(), args.task, task)
        dry_run_path = run_dir / "writeback_dry_run.json"
        dry_run_path.write_text(json.dumps(dry_run_result, ensure_ascii=False, indent=2), encoding="utf-8")
        result["steps"]["dry_run"] = dry_run_result
        result["artifacts"]["dry_run"] = str(dry_run_path)
        if dry_run_result.get("status") != "ok":
            result["status"] = "error"

    if args.diff:
        if not args.original_file or not args.candidate_file:
            raise SmartbiError("writeback --diff requires --original-file and --candidate-file", code="config_error")
        from diff_smartbi_writeback_workbooks import build_diff
        from validate_smartbi_writeback_input import get_task as get_writeback_task
        from validate_smartbi_writeback_input import load_config as load_writeback_config

        config = load_writeback_config(config_path)
        task = get_writeback_task(config, args.task)
        diff_result = build_diff(
            original=Path(args.original_file).expanduser(),
            candidate=Path(args.candidate_file).expanduser(),
            task_name=args.task,
            task=task,
            max_changed_rows=args.max_changed_rows,
        )
        diff_path = run_dir / "writeback_diff.json"
        diff_path.write_text(json.dumps(diff_result, ensure_ascii=False, indent=2), encoding="utf-8")
        result["steps"]["diff"] = diff_result
        result["artifacts"]["diff"] = str(diff_path)
        if diff_result.get("status") != "ok":
            result["status"] = "error"

    if args.http_probe:
        http_probe_result = run_writeback_http_probe(
            config_path=config_path,
            task_name=args.task,
            username=args.username,
            password=args.password,
            input_filename=args.file or args.candidate_file,
        )
        http_probe_path = run_dir / "writeback_http_probe.json"
        http_probe_path.write_text(json.dumps(http_probe_result, ensure_ascii=False, indent=2), encoding="utf-8")
        result["steps"]["http_probe"] = http_probe_result
        result["artifacts"]["http_probe"] = str(http_probe_path)
        if http_probe_result.get("status") != "ok":
            result["status"] = "error"

    if args.probe:
        if args.username:
            os.environ["SMARTBI_USERNAME"] = str(args.username)
        if args.password:
            os.environ["SMARTBI_PASSWORD"] = str(args.password)
        from probe_smartbi_writeback_target import run_probe as run_writeback_probe

        probe_result = asyncio.run(
            run_writeback_probe(
                config_path=config_path,
                task_name=args.task,
                out_root=run_dir / "target_probe",
                headless=not args.headed,
                browser_channel=args.browser_channel,
            )
        )
        result["steps"]["probe"] = probe_result
        result["artifacts"]["probe"] = (probe_result.get("artifacts") or {}).get("result")
        if probe_result.get("status") != "ok":
            result["status"] = "error"

    if args.shadow_execute:
        shadow_plan = run_writeback_shadow_execute(
            config_path=config_path,
            task_name=args.task,
            username=args.username,
            password=args.password,
            candidate_file=Path(args.candidate_file).expanduser() if args.candidate_file else None,
            rollback_file=Path(args.rollback_file).expanduser() if args.rollback_file else None,
            expected_diff_file=Path(args.expected_diff).expanduser() if args.expected_diff else None,
            max_changed_rows=args.max_changed_rows,
            run_dir=run_dir,
        )
        result["steps"]["shadow_execute"] = shadow_plan
        result["artifacts"]["shadow_execution_plan"] = shadow_plan.get("artifact")
        if shadow_plan.get("status") != "ok":
            result["status"] = "error"

    summary_path = run_dir / "writeback_run.json"
    summary_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    result["artifacts"]["summary"] = str(summary_path)
    return result


def require_auth(username: str | None, password: str | None) -> None:
    if not username or not password:
        raise SmartbiError("SMARTBI_USERNAME/SMARTBI_PASSWORD or --username/--password is required", code="auth_error")


def login_client(username: str | None, password: str | None) -> SmartbiClient:
    require_auth(username, password)
    client = SmartbiClient()
    client.login(str(username), str(password))
    return client


def normalize_catalog_element(element: dict[str, Any], path: str) -> dict[str, Any]:
    return {
        "id": element.get("id"),
        "alias": element.get("alias") or element.get("name"),
        "name": element.get("name"),
        "type": element.get("type"),
        "has_child": bool(element.get("hasChild")),
        "path": path,
    }


def resolve_catalog_path(client: SmartbiClient, path: str) -> tuple[str, list[dict[str, Any]], str]:
    parts = [part for part in path.strip("/").split("/") if part]
    if not parts:
        raise SmartbiError("Catalog path is required")
    roots = client.root_elements()
    root = find_catalog_element(roots, parts[0])
    current_id = required_element_id(root)
    resolved = [normalize_catalog_element(root, parts[0])]
    current_path = parts[0]
    for part in parts[1:]:
        children = client.child_elements(current_id)
        child = find_catalog_element(children, part)
        current_id = required_element_id(child)
        current_path = f"{current_path}/{part}"
        resolved.append(normalize_catalog_element(child, current_path))
    return current_id, resolved, current_path


def find_catalog_element(elements: list[dict[str, Any]], label: str) -> dict[str, Any]:
    for element in elements:
        if label in {str(element.get("alias") or ""), str(element.get("name") or ""), str(element.get("id") or "")}:
            return element
    available = ", ".join(sorted(str(element.get("alias") or element.get("name") or element.get("id")) for element in elements))
    raise SmartbiError(f"Cannot find catalog path segment '{label}'. Available: {available}")


def required_element_id(element: dict[str, Any]) -> str:
    element_id = element.get("id")
    if not isinstance(element_id, str) or not element_id:
        raise SmartbiError("Catalog element missing id")
    return element_id


def list_catalog(
    client: SmartbiClient,
    path: str,
    recursive: bool = False,
    max_depth: int = 2,
) -> dict[str, Any]:
    parent_id, resolved, resolved_path = resolve_catalog_path(client, path)
    children = collect_catalog_children(client, parent_id, resolved_path, recursive=recursive, max_depth=max_depth)
    return {
        "path": resolved_path,
        "resolved": resolved,
        "children": children,
    }


def collect_catalog_children(
    client: SmartbiClient,
    parent_id: str,
    parent_path: str,
    recursive: bool,
    max_depth: int,
    depth: int = 1,
) -> list[dict[str, Any]]:
    children = []
    for child in client.child_elements(parent_id):
        label = str(child.get("alias") or child.get("name") or child.get("id"))
        child_path = f"{parent_path}/{label}"
        normalized = normalize_catalog_element(child, child_path)
        children.append(normalized)
        if recursive and normalized["has_child"] and depth < max_depth:
            children.extend(
                collect_catalog_children(
                    client,
                    required_element_id(child),
                    child_path,
                    recursive=recursive,
                    max_depth=max_depth,
                    depth=depth + 1,
                )
            )
    return children


def inspect_report(
    client: SmartbiClient,
    report_id: str,
    report_path: str | None = None,
    task_name: str | None = None,
) -> dict[str, Any]:
    context = client.open_report_context(report_id)
    params = params_from_context(context)
    sheets = context.get("visibleSheetNames") or []
    alias = context.get("alias") or context.get("name") or report_id
    parameter_summaries = [summarize_param(param) for param in params]
    return {
        "report_id": report_id,
        "alias": alias,
        "name": context.get("name"),
        "report_path": report_path,
        "visible_sheets": sheets,
        "parameter_count": len(parameter_summaries),
        "parameters": parameter_summaries,
        "task_draft": build_task_draft(
            report_id=report_id,
            report_path=report_path,
            alias=str(alias),
            task_name=task_name,
        ),
    }


def draft_catalog_config(
    client: SmartbiClient,
    path: str,
    recursive: bool,
    max_depth: int,
    max_reports: int | None,
    out_path: Path | None,
) -> dict[str, Any]:
    catalog = list_catalog(client, path, recursive=recursive, max_depth=max_depth)
    spreadsheet_reports = [
        child for child in catalog["children"]
        if child.get("type") == "SPREADSHEET_REPORT"
    ]
    if max_reports is not None:
        spreadsheet_reports = spreadsheet_reports[:max_reports]

    tasks: dict[str, Any] = {}
    inspections: dict[str, Any] = {}
    for report in spreadsheet_reports:
        report_id = str(report["id"])
        report_path = str(report["path"])
        alias = str(report.get("alias") or report.get("name") or report_id)
        task_name = task_name_for_report(alias, report_id)
        inspection = inspect_report(client, report_id, report_path=report_path, task_name=task_name)
        tasks.update(inspection["task_draft"])
        inspections[task_name] = {
            "report_id": report_id,
            "report_path": report_path,
            "alias": alias,
            "visible_sheets": inspection["visible_sheets"],
            "parameter_count": inspection["parameter_count"],
            "parameters": inspection["parameters"],
        }

    draft = {
        "version": 1,
        "source_catalog": catalog["path"],
        "report_count": len(spreadsheet_reports),
        "tasks": tasks,
        "inspections": inspections,
    }
    if out_path is not None:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json.dumps(draft, ensure_ascii=False, indent=2), encoding="utf-8")
        draft["output"] = str(out_path)
    return draft


def summarize_param(param: dict[str, Any]) -> dict[str, Any]:
    return {
        "key": param.get("alias") or param.get("name") or param.get("id"),
        "id": param.get("id"),
        "name": param.get("name"),
        "alias": param.get("alias"),
        "value": param.get("value"),
        "displayValue": param.get("displayValue"),
    }


def build_task_draft(
    report_id: str,
    report_path: str | None,
    alias: str,
    task_name: str | None,
) -> dict[str, Any]:
    effective_task_name = task_name or task_name_for_report(alias, report_id)
    return {
        effective_task_name: {
            "enabled": True,
            "description": f"{alias}，配置草稿。",
            "report": {
                "id": report_id,
                "path": report_path or "",
                "type": "SPREADSHEET_REPORT",
            },
            "filters": {
                "mode": "default"
            },
            "output": {
                "type": "file",
                "dir": "outputs/bi_exports/{task}/{run_date}",
            },
        }
    }


def task_name_for_report(alias: str, report_id: str) -> str:
    slug = slugify(alias)
    if slug != "smartbi_report":
        return slug
    return f"smartbi_{report_id[-8:].lower()}"


def slugify(value: str) -> str:
    slug = re.sub(r"[^0-9A-Za-z_]+", "_", value).strip("_").lower()
    return slug or "smartbi_report"


def extract_js_object(text: str, marker: str) -> str:
    start = text.find(marker)
    if start < 0:
        raise SmartbiError(f"Cannot find marker: {marker}")
    brace = text.find("{", start + len(marker))
    if brace < 0:
        raise SmartbiError(f"Cannot find object after marker: {marker}")
    depth = 0
    in_string = False
    escaped = False
    for index in range(brace, len(text)):
        char = text[index]
        if in_string:
            if escaped:
                escaped = False
            elif char == "\\":
                escaped = True
            elif char == '"':
                in_string = False
        else:
            if char == '"':
                in_string = True
            elif char == "{":
                depth += 1
            elif char == "}":
                depth -= 1
                if depth == 0:
                    return text[brace : index + 1]
    raise SmartbiError(f"Unterminated object after marker: {marker}")


def default_filename(content_disposition: str | None) -> str | None:
    if not content_disposition:
        return None
    match = re.search(r"filename\*?=(?:UTF-8'')?\"?([^\";]+)\"?", content_disposition)
    if not match:
        return None
    return urllib.parse.unquote(match.group(1))


def required_str(data: dict[str, object], key: str) -> str:
    value = data.get(key)
    if not isinstance(value, str) or not value:
        raise SmartbiError(f"Report context missing required field: {key}")
    return value


def unique_path(path: Path, overwrite: bool) -> Path:
    if overwrite or not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    for index in range(1, 1000):
        candidate = path.with_name(f"{stem} ({index}){suffix}")
        if not candidate.exists():
            return candidate
    raise SmartbiError(f"Too many existing files matching {path.name}")


def doctor_check(
    name: str,
    ok: bool,
    *,
    code: str | None = None,
    message: str | None = None,
    required: bool = True,
    details: dict[str, Any] | None = None,
) -> dict[str, Any]:
    check: dict[str, Any] = {
        "name": name,
        "ok": ok,
        "required": required,
    }
    if code:
        check["code"] = code
    if message:
        check["message"] = message
    if details:
        check["details"] = details
    return check


def doctor_error_check(name: str, error: Exception, *, required: bool = True) -> dict[str, Any]:
    code = error.code if isinstance(error, SmartbiError) else "doctor_check_failed"
    return doctor_check(name, False, code=code, message=str(error), required=required)


def load_json_object(path: Path, code: str = "config_error") -> dict[str, Any]:
    try:
        with path.open("r", encoding="utf-8") as handle:
            data = json.load(handle)
    except FileNotFoundError as error:
        raise SmartbiError(f"JSON file not found: {path}", code=code) from error
    except json.JSONDecodeError as error:
        raise SmartbiError(f"Invalid JSON file {path}: {error}", code=code) from error
    if not isinstance(data, dict):
        raise SmartbiError(f"JSON file must contain an object: {path}", code=code)
    return data


def check_no_obvious_plaintext_credentials() -> dict[str, Any]:
    targets = [ROOT / "scripts", ROOT / "configs", ROOT / "docs", ROOT / "recipes"]
    findings: list[str] = []
    assignment_pattern = re.compile(
        r"(SMARTBI_PASSWORD|password)\s*[:=]\s*['\"](?!\.\.\.|<|\$?\{?SMARTBI_PASSWORD)([^'\"]{6,})['\"]",
        re.IGNORECASE,
    )
    token_patterns = [
        re.compile(r"sk-[A-Za-z0-9_-]{20,}"),
        re.compile(r"xox[baprs]-[A-Za-z0-9-]{20,}"),
        re.compile(r"AKIA[0-9A-Z]{16}"),
        re.compile(r"ASIA[0-9A-Z]{16}"),
        re.compile(r"-----BEGIN (?:RSA |OPENSSH )?PRIVATE KEY-----"),
    ]
    for target in targets:
        if not target.exists():
            continue
        for path in target.rglob("*"):
            if not path.is_file():
                continue
            try:
                text = path.read_text(encoding="utf-8")
            except UnicodeDecodeError:
                continue
            if assignment_pattern.search(text):
                findings.append(f"{path.relative_to(ROOT)}: possible plaintext password assignment")
            for pattern in token_patterns:
                if pattern.search(text):
                    findings.append(f"{path.relative_to(ROOT)}: possible secret token")
    return doctor_check(
        "no_obvious_plaintext_credentials",
        not findings,
        code=None if not findings else "plaintext_credentials_found",
        message=None if not findings else "Possible credentials or tokens found in project text files",
        details={"findings": findings[:20], "truncated": len(findings) > 20} if findings else None,
    )


def run_doctor(args: argparse.Namespace) -> dict[str, Any]:
    checks: list[dict[str, Any]] = []
    read_config_path = Path(args.config).expanduser()
    writeback_config_path = Path(args.writeback_config).expanduser()

    checks.append(
        doctor_check(
            "python_version",
            sys.version_info >= (3, 10),
            code=None if sys.version_info >= (3, 10) else "python_version_unsupported",
            message=None if sys.version_info >= (3, 10) else "Python 3.10+ is recommended for teammate runs",
            details={"version": sys.version.split()[0]},
        )
    )
    checks.append(
        doctor_check(
            "openpyxl_available",
            importlib.util.find_spec("openpyxl") is not None,
            code="dependency_missing" if importlib.util.find_spec("openpyxl") is None else None,
            message="openpyxl is required for workbook validation" if importlib.util.find_spec("openpyxl") is None else None,
        )
    )
    checks.append(
        doctor_check(
            "pandas_available_for_inspection",
            importlib.util.find_spec("pandas") is not None,
            code="optional_dependency_missing" if importlib.util.find_spec("pandas") is None else None,
            message="pandas is recommended for inspect_smartbi_workbook.py" if importlib.util.find_spec("pandas") is None else None,
            required=False,
        )
    )

    try:
        read_config = load_config(read_config_path)
        checks.append(
            doctor_check(
                "read_config_loadable",
                True,
                details={"path": str(read_config_path), "task_count": len(read_config.get("tasks", {}))},
            )
        )
    except Exception as error:
        checks.append(doctor_error_check("read_config_loadable", error))

    try:
        writeback_config = load_json_object(writeback_config_path)
        tasks = writeback_config.get("tasks")
        if not isinstance(tasks, dict) or not tasks:
            raise SmartbiError("Writeback config must include a non-empty tasks object", code="config_error")
        checks.append(
            doctor_check(
                "writeback_config_loadable",
                True,
                details={"path": str(writeback_config_path), "task_count": len(tasks)},
            )
        )
    except Exception as error:
        checks.append(doctor_error_check("writeback_config_loadable", error))

    required_docs = [
        ROOT / "docs" / "error_codes.md",
        ROOT / "docs" / "security_boundary.md",
        ROOT / "docs" / "smartbi_writeback_devkit_operator_guide.md",
        ROOT / "recipes" / "new_report_onboarding.md",
        ROOT / "recipes" / "read_only_export_chain.md",
        ROOT / "recipes" / "writeback_shadow_validation.md",
    ]
    missing_docs = [str(path.relative_to(ROOT)) for path in required_docs if not path.exists()]
    checks.append(
        doctor_check(
            "teammate_docs_present",
            not missing_docs,
            code=None if not missing_docs else "teammate_docs_missing",
            message=None if not missing_docs else "Required teammate handoff docs are missing",
            details={"missing": missing_docs} if missing_docs else None,
        )
    )
    checks.append(check_no_obvious_plaintext_credentials())

    username = args.username
    password = args.password
    credentials_present = bool(username and password)
    checks.append(
        doctor_check(
            "credentials_present",
            credentials_present,
            code=None if credentials_present else "credentials_missing",
            message=None if credentials_present else "SMARTBI_USERNAME/SMARTBI_PASSWORD or --username/--password is required for online checks",
            required=bool(args.online),
        )
    )

    if args.online and credentials_present:
        try:
            client = SmartbiClient()
            client.request("index.jsp?time=1778858593300")
            checks.append(doctor_check("base_url_reachable", True, details={"base_url": client.base_url}))
            client.login(str(username), str(password))
            checks.append(doctor_check("login_success", True))
            root_elements = client.root_elements()
            checks.append(
                doctor_check(
                    "catalog_root_readable",
                    True,
                    details={"root_count": len(root_elements)},
                )
            )
            if args.catalog_path:
                _, elements, _ = resolve_catalog_path(client, args.catalog_path)
                checks.append(
                    doctor_check(
                        "catalog_path_readable",
                        True,
                        details={"path": args.catalog_path, "child_count": len(elements)},
                    )
                )
        except Exception as error:
            checks.append(doctor_error_check("online_smartbi_check", error))
    elif args.online:
        checks.append(
            doctor_check(
                "online_smartbi_check",
                False,
                code="credentials_missing",
                message="Skipped online SmartBI checks because credentials are missing",
            )
        )
    else:
        checks.append(
            doctor_check(
                "online_smartbi_check",
                True,
                code="skipped",
                message="Skipped by default. Rerun with --online to test SmartBI login and catalog read permission.",
                required=False,
            )
        )

    failed_required = [check for check in checks if check.get("required", True) and not check.get("ok")]
    if failed_required:
        first = failed_required[0]
        next_action = first.get("message") or f"Fix {first.get('name')} before handing this CLI to a teammate."
    elif not args.online:
        next_action = "Offline checks passed. Run doctor --online --json after credentials are injected to verify SmartBI access."
    else:
        next_action = "Doctor checks passed. Use recipes/ for the teammate workflow and preserve run artifacts."
    return {
        "status": "ok" if not failed_required else "failed",
        "mode": "online" if args.online else "offline",
        "boundary": {
            "smartbi_login": bool(args.online and credentials_present),
            "excel_export": False,
            "writeback_upload": False,
            "external_write": False,
            "credentials_logged": False,
        },
        "checks": checks,
        "next_action": next_action,
    }


def build_parser() -> argparse.ArgumentParser:
    load_local_env()
    parser = argparse.ArgumentParser(description="SmartBI Data CLI")
    subparsers = parser.add_subparsers(dest="command", required=True)

    doctor_parser = subparsers.add_parser(
        "doctor",
        help="Run teammate preflight checks. Offline by default; --online tests SmartBI login/catalog only.",
    )
    add_auth_args(doctor_parser)
    doctor_parser.add_argument("--config", default=str(DEFAULT_CONFIG))
    doctor_parser.add_argument("--writeback-config", default=str(DEFAULT_WRITEBACK_CONFIG))
    doctor_parser.add_argument("--online", action="store_true", help="Login and read SmartBI catalog root; no export or upload.")
    doctor_parser.add_argument("--catalog-path", help="Optional catalog path to resolve during --online checks.")
    doctor_parser.add_argument("--json", action="store_true", help="Print machine-readable result.")

    export_parser = subparsers.add_parser(
        "export-outbound-quality",
        help="Export 海外产运/外呼/益智外呼质量监控 as Excel with default filters.",
    )
    add_auth_args(export_parser)
    export_parser.add_argument("--out-dir", default=str(Path.home() / "Desktop"))
    export_parser.add_argument("--overwrite", action="store_true")
    export_parser.add_argument("--json", action="store_true", help="Print machine-readable result.")

    run_parser = subparsers.add_parser("run", help="Run a config-defined Smartbi export task.")
    add_auth_args(run_parser)
    run_parser.add_argument("--config", default=str(DEFAULT_CONFIG))
    run_parser.add_argument("--task", required=True)
    run_parser.add_argument("--run-date", help="Anchor date for date_window and {run_date}, YYYY-MM-DD. Defaults to today.")
    run_parser.add_argument("--dry-run", action="store_true")
    run_parser.add_argument("--overwrite", action="store_true")
    run_parser.add_argument("--json", action="store_true", help="Print machine-readable result.")

    catalog_parser = subparsers.add_parser("catalog-list", help="List a Smartbi catalog path.")
    add_auth_args(catalog_parser)
    catalog_parser.add_argument("--path", required=True, help="Catalog path, e.g. 分析报表/海外直播业务线/海外产运/外呼")
    catalog_parser.add_argument("--recursive", action="store_true")
    catalog_parser.add_argument("--max-depth", type=int, default=2)
    catalog_parser.add_argument("--json", action="store_true", help="Print machine-readable result.")

    inspect_parser = subparsers.add_parser("inspect-report", help="Inspect a SmartBI report and print parameter metadata.")
    add_auth_args(inspect_parser)
    inspect_parser.add_argument("--report-id", required=True)
    inspect_parser.add_argument("--report-path")
    inspect_parser.add_argument("--task-name")
    inspect_parser.add_argument("--json", action="store_true", help="Print machine-readable result.")

    draft_parser = subparsers.add_parser("catalog-draft", help="Generate config draft for spreadsheet reports in a catalog path.")
    add_auth_args(draft_parser)
    draft_parser.add_argument("--path", required=True, help="Catalog path to inspect.")
    draft_parser.add_argument("--recursive", action="store_true")
    draft_parser.add_argument("--max-depth", type=int, default=2)
    draft_parser.add_argument("--max-reports", type=int)
    draft_parser.add_argument("--out", help="Optional JSON output path.")
    draft_parser.add_argument("--json", action="store_true", help="Print machine-readable result.")

    matrix_parser = subparsers.add_parser(
        "route-matrix",
        help="Run offline BI route matrix cases without logging into SmartBI or exporting Excel.",
    )
    matrix_parser.add_argument("--config", default="configs/bi_route_matrix_p0.json")
    matrix_parser.add_argument("--run-id")
    matrix_parser.add_argument("--case-id", action="append", default=[])
    matrix_parser.add_argument("--limit", type=int, default=5)
    matrix_parser.add_argument("--json", action="store_true", help="Print machine-readable result.")

    check_parser = subparsers.add_parser(
        "writeback-check",
        help="Locally inspect a writeback workbook; no SmartBI login and no upload.",
    )
    check_parser.add_argument("--config", default=str(DEFAULT_WRITEBACK_CONFIG))
    check_parser.add_argument("--task", required=True)
    check_parser.add_argument("--file", required=True)
    check_parser.add_argument("--operator")
    check_parser.add_argument("--out")
    check_parser.add_argument("--json", action="store_true", help="Print machine-readable result.")

    upload_parser = subparsers.add_parser(
        "writeback-upload",
        help="Developer/operator alias for one confirmed SmartBI writeback upload plus post-verify.",
    )
    add_auth_args(upload_parser)
    upload_parser.add_argument("--config", default=str(DEFAULT_WRITEBACK_CONFIG))
    upload_parser.add_argument("--task", required=True)
    upload_parser.add_argument("--file", required=True)
    upload_parser.add_argument("--confirm", action="store_true", help="Required: sends one DataAcquisitionServlet upload.")
    upload_parser.add_argument("--operator")
    upload_parser.add_argument("--run-id")
    upload_parser.add_argument("--out-root", default="outputs/smartbi_writeback_cli")
    upload_parser.add_argument("--json", action="store_true", help="Print machine-readable result.")

    diagnose_parser = subparsers.add_parser(
        "writeback-diagnose",
        help="Summarize a saved writeback run directory; no SmartBI login and no upload.",
    )
    diagnose_parser.add_argument("--run-dir", required=True)
    diagnose_parser.add_argument("--out")
    diagnose_parser.add_argument("--json", action="store_true", help="Print machine-readable result.")

    onboard_parser = subparsers.add_parser(
        "writeback-onboard",
        help="Generate a task config draft and onboarding report for a writeback table.",
    )
    add_auth_args(onboard_parser)
    onboard_parser.add_argument("--target-url")
    onboard_parser.add_argument("--target-path")
    onboard_parser.add_argument("--sample-file", required=True)
    onboard_parser.add_argument("--task", required=True)
    onboard_parser.add_argument("--post-verify-url")
    onboard_parser.add_argument("--post-verify-path")
    onboard_parser.add_argument("--run-id")
    onboard_parser.add_argument("--out-root", default="outputs/smartbi_writeback_cli/onboarding")
    onboard_parser.add_argument("--offline", action="store_true", help="Generate draft artifacts without SmartBI login/probe/template download.")
    onboard_parser.add_argument("--json", action="store_true", help="Print machine-readable result.")

    writeback_parser = subparsers.add_parser(
        "writeback",
        help="Validate or probe a SmartBI Excel writeback target without uploading.",
    )
    add_auth_args(writeback_parser)
    writeback_parser.add_argument("--config", default=str(DEFAULT_WRITEBACK_CONFIG))
    writeback_parser.add_argument("--task", required=True)
    writeback_parser.add_argument("--file", help="Input workbook for --dry-run.")
    writeback_parser.add_argument("--dry-run", action="store_true", help="Validate local workbook only; no SmartBI login.")
    writeback_parser.add_argument("--diff", action="store_true", help="Compare original and candidate workbooks without SmartBI login.")
    writeback_parser.add_argument("--original-file", help="Rollback/source workbook for --diff.")
    writeback_parser.add_argument("--candidate-file", help="Human-edited test workbook for --diff.")
    writeback_parser.add_argument("--rollback-file", help="Original rollback workbook required for --shadow-execute.")
    writeback_parser.add_argument("--expected-diff", help="Expected diff manifest required for --shadow-execute.")
    writeback_parser.add_argument("--shadow-execute", action="store_true", help="Build audited execution plan without submitting upload.")
    writeback_parser.add_argument("--max-changed-rows", type=int, default=3)
    writeback_parser.add_argument("--http-probe", action="store_true", help="Login through SmartbiClient and probe servlet endpoints without browser or upload.")
    writeback_parser.add_argument("--probe", action="store_true", help="Login and open target page; no file selection or upload.")
    writeback_parser.add_argument("--single-upload", action="store_true", help="Owner-approved single DataAcquisitionServlet upload with parameter-panel initialization.")
    writeback_parser.add_argument("--execute", action="store_true", help="Blocked in this slice; reserved for a future confirmed upload.")
    writeback_parser.add_argument("--confirm-writeback", action="store_true", help="Reserved for future real upload confirmation.")
    writeback_parser.add_argument("--owner-approval-token", help="Run-id scoped owner approval token required for guarded execute.")
    writeback_parser.add_argument("--upload-window-start", help="ISO datetime start for guarded execute upload window.")
    writeback_parser.add_argument("--upload-window-end", help="ISO datetime end for guarded execute upload window.")
    writeback_parser.add_argument("--candidate-sha256", help="Expected SHA-256 for --candidate-file in guarded execute.")
    writeback_parser.add_argument("--rollback-sha256", help="Expected SHA-256 for --rollback-file in guarded execute.")
    writeback_parser.add_argument("--expected-diff-sha256", help="Expected SHA-256 for --expected-diff in guarded execute.")
    writeback_parser.add_argument("--run-id")
    writeback_parser.add_argument("--out-root", default="outputs/smartbi_writeback_cli")
    writeback_parser.add_argument("--headed", action="store_true")
    writeback_parser.add_argument("--browser-channel", default="chrome")
    writeback_parser.add_argument("--json", action="store_true", help="Print machine-readable result.")
    return parser


def add_auth_args(parser: argparse.ArgumentParser) -> None:
    parser.add_argument("--username", default=os.environ.get("SMARTBI_USERNAME"))
    parser.add_argument("--password", default=os.environ.get("SMARTBI_PASSWORD"))


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    try:
        if args.command == "doctor":
            result = run_doctor(args)
        elif args.command == "export-outbound-quality":
            result = export_outbound_quality(args)
        elif args.command == "run":
            config_path = Path(args.config).expanduser()
            config = load_config(config_path)
            task = get_task(config, args.task)
            result = run_task(
                args.task,
                task,
                config_path,
                args.username,
                args.password,
                args.dry_run,
                args.overwrite,
                dt.date.fromisoformat(args.run_date) if args.run_date else None,
            )
        elif args.command == "catalog-list":
            client = login_client(args.username, args.password)
            result = list_catalog(client, args.path, recursive=args.recursive, max_depth=args.max_depth)
        elif args.command == "inspect-report":
            client = login_client(args.username, args.password)
            result = inspect_report(client, args.report_id, report_path=args.report_path, task_name=args.task_name)
        elif args.command == "catalog-draft":
            client = login_client(args.username, args.password)
            out_path = Path(args.out).expanduser() if args.out else None
            result = draft_catalog_config(
                client,
                args.path,
                recursive=args.recursive,
                max_depth=args.max_depth,
                max_reports=args.max_reports,
                out_path=out_path,
            )
        elif args.command == "route-matrix":
            result = run_route_matrix(args)
        elif args.command == "writeback-check":
            result = run_writeback_check(args)
        elif args.command == "writeback-upload":
            result = run_writeback_upload(args)
        elif args.command == "writeback-diagnose":
            result = run_writeback_diagnose(args)
        elif args.command == "writeback-onboard":
            result = run_writeback_onboard(args)
        elif args.command == "writeback":
            result = run_writeback_command(args)
        else:
            raise SmartbiError(f"Unsupported command: {args.command}")
    except SmartbiError as error:
        payload = {"status": "error", "error": {"code": error.code, "message": str(error)}}
        if getattr(args, "json", False):
            print(json.dumps(payload, ensure_ascii=False, indent=2), file=sys.stderr)
        else:
            print(f"{error.code}: {error}", file=sys.stderr)
        return 2

    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        print_result(result)
    return 2 if isinstance(result, dict) and result.get("status") in {"error", "failed"} else 0


def run_route_matrix(args: argparse.Namespace) -> dict[str, Any]:
    import run_bi_route_matrix

    config_path = Path(args.config).expanduser()
    if not config_path.is_absolute():
        config_path = ROOT / config_path
    config = run_bi_route_matrix.load_config(config_path)
    registry = run_bi_route_matrix.resolve_path(
        config.get("registry"),
        ROOT / "outputs" / "bi_catalog_registry" / "bi_report_route_index_current.json",
    )
    output_root = run_bi_route_matrix.resolve_path(
        config.get("output_root"),
        ROOT / "outputs" / "bi_catalog_registry" / "matrix_runs",
    )
    run_id = args.run_id or dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    run_root = output_root / run_id
    run_root.mkdir(parents=True, exist_ok=True)

    wanted = {item for value in args.case_id for item in str(value).split(",") if item}
    cases = [case for case in config["cases"] if not wanted or str(case.get("id")) in wanted]
    if not cases:
        raise SmartbiError("No route matrix cases selected", code="config_error")

    results = []
    command_results = []
    for case in cases:
        packet, command_result = run_bi_route_matrix.run_packet(case, registry, run_root, args.limit)
        command_results.append(command_result)
        results.append(run_bi_route_matrix.evaluate_case(case, packet, command_result))

    summary_json, summary_md = run_bi_route_matrix.write_summary(run_root, config_path, results, command_results)
    counts = {
        "total": len(results),
        "pass": sum(1 for item in results if item["status"] == "pass"),
        "weak": sum(1 for item in results if item["status"] == "weak"),
        "fail": sum(1 for item in results if item["status"] == "fail"),
    }
    return {
        "status": "ok" if counts["fail"] == 0 else "error",
        "mode": "offline_route_matrix",
        "boundary": {
            "no_smartbi_login": True,
            "no_excel_export": True,
            "no_credentials_required": True,
        },
        "summary_json": str(summary_json),
        "summary_md": str(summary_md),
        "counts": counts,
    }


def export_outbound_quality(args: argparse.Namespace) -> dict[str, Any]:
    client = login_client(args.username, args.password)
    context = client.open_report_context(REPORT_ID)
    filename, body = client.export_spreadsheet_report(REPORT_ID, context)
    out_dir = Path(args.out_dir).expanduser()
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = unique_path(out_dir / filename, args.overwrite)
    out_path.write_bytes(body)
    validation = validate_xlsx(out_path)
    return {
        "status": "exported",
        "report_path": REPORT_PATH,
        "report_id": REPORT_ID,
        "output": str(out_path),
        "bytes": validation["bytes"],
        "sheets": validation["sheets"],
        "default_filename": filename,
    }


def print_result(result: dict[str, Any]) -> None:
    status = result.get("status")
    if status == "dry_run":
        print("Dry run plan:")
        print(json.dumps(result["plan"], ensure_ascii=False, indent=2))
        return
    if "children" in result:
        print(f"Catalog: {result['path']}")
        for child in result["children"]:
            child_type = child.get("type") or ""
            marker = "/" if child.get("has_child") else ""
            print(f"{child_type}\t{child.get('alias')}{marker}\t{child.get('id')}")
        return
    if "parameters" in result:
        print(f"Report: {result.get('alias')} ({result.get('report_id')})")
        print(f"Sheets: {', '.join(result.get('visible_sheets') or [])}")
        print("Parameters:")
        for param in result["parameters"]:
            print(f"- {param.get('key')}: value={param.get('value')} display={param.get('displayValue')}")
        return
    if "inspections" in result:
        print(f"Catalog draft: {result.get('source_catalog')}")
        print(f"Spreadsheet reports: {result.get('report_count')}")
        if result.get("output"):
            print(f"Output: {result['output']}")
        for task_name, inspection in result["inspections"].items():
            print(f"- {task_name}: {inspection.get('alias')} ({inspection.get('parameter_count')} parameters)")
        return
    print(f"Status: {status}")
    if result.get("report_path"):
        print(f"Report: {result['report_path']}")
    if result.get("output"):
        print(f"Output: {result['output']}")
    if result.get("bytes"):
        print(f"Bytes: {result['bytes']}")
    if result.get("sheets"):
        print(f"Sheets: {', '.join(result['sheets'])}")


if __name__ == "__main__":
    raise SystemExit(main())
